#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK_HEADERS = {
    "review_status": "검토상태",
    "export_flag": "내보내기",
    "product_id": "productId",
    "product_name": "상품명",
    "sku_code": "SKU",
    "branduid": "branduid",
    "sabang_code": "사방넷코드",
    "category": "카테고리",
    "base_profile_id": "기본프로필",
    "procurement_type": "조달유형",
    "source_cost_type": "원가타입",
    "source_cost_seed": "시드원가",
    "source_cost_input": "확정원가입력",
    "exchange_rate": "환율",
    "selling_price_seed": "시드판매가",
    "selling_price_input": "현재판매가입력",
    "branch_management_rate": "지사관리비율",
    "freight_per_unit": "단위당운임",
    "tariff_rate": "관세율",
    "packaging_cost": "포장비",
    "inspection_cost": "검품비",
    "loss_rate": "손실충당율",
    "vat_rate": "VAT율",
    "include_vat_in_cogs": "VAT포함",
    "shipping_per_unit": "단위당배송비",
    "source_confidence": "원가신뢰도",
    "verified_by": "검증자",
    "verified_at": "검증일",
    "review_note": "검토메모",
    "seed_cost_history": "시드원가이력",
    "alias_summary": "별칭/원본명",
}
FIELD_COLUMNS = {
    "branchManagementRate": "branch_management_rate",
    "freightPerUnit": "freight_per_unit",
    "tariffRate": "tariff_rate",
    "packagingCost": "packaging_cost",
    "inspectionCost": "inspection_cost",
    "lossRate": "loss_rate",
    "vatRate": "vat_rate",
    "includeVatInCogs": "include_vat_in_cogs",
    "shippingPerUnit": "shipping_per_unit",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export reviewed PRESSCO21 cost ledger workbook to OpenClaw JSON.")
    parser.add_argument("--input", required=True, help="Path to ledger XLSX")
    parser.add_argument("--output", required=True, help="Path to reviewed JSON output")
    parser.add_argument("--seed", default="", help="Optional seed JSON to preserve source snapshot metadata")
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    try:
        numeric = float(value)
    except Exception:
        text = normalize_text(value).replace(",", "")
        if not text:
            return None
        try:
            numeric = float(text)
        except Exception:
            return None
    if numeric.is_integer():
        return int(numeric)
    return round(numeric, 4)


def parse_flag(value: Any) -> bool:
    text = normalize_text(value).lower()
    return text in {"y", "yes", "true", "1"}


def split_aliases(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def parse_history(value: Any) -> list[float | int]:
    text = normalize_text(value)
    if not text:
        return []
    parts = [part.strip() for part in text.replace("[", "").replace("]", "").split(",") if part.strip()]
    values = []
    for part in parts:
        numeric = parse_number(part)
        if numeric is not None:
            values.append(numeric)
    return values


def unique(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = normalize_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def build_sheet_index(ws) -> dict[str, int]:
    index: dict[str, int] = {}
    for column_index, cell in enumerate(ws[1], start=1):
        label = normalize_text(cell.value)
        if label:
            index[label] = column_index
    return index


def row_value(row: tuple[Any, ...], index: dict[str, int], header_key: str) -> Any:
    label = WORKBOOK_HEADERS[header_key]
    column_index = index.get(label)
    if not column_index:
        return None
    zero_based = column_index - 1
    if zero_based >= len(row):
        return None
    return row[zero_based]


def load_seed(seed_path: Path) -> tuple[dict[str, Any], dict[str, Any], dict[tuple[str, str], dict[str, Any]]]:
    if not seed_path.exists():
        return {}, {}, {}
    payload = json.loads(seed_path.read_text(encoding="utf-8"))
    by_id = {}
    by_identity = {}
    for product in payload.get("products", []):
        product_id = normalize_text(product.get("productId"))
        if product_id:
            by_id[product_id] = product
        identity = (
            normalize_text((product.get("branduids") or [""])[0]),
            normalize_text(product.get("productName")),
        )
        if any(identity):
            by_identity[identity] = product
    return payload, by_id, by_identity


def status_from_review(review_status: str) -> str:
    mapping = {
        "seeded_review_pending": "draft_seeded_from_master",
        "conflict_needs_review": "draft_conflict_from_master",
        "review_in_progress": "review_in_progress",
        "verified_ready": "verified_ready",
        "new_draft": "draft_new_product",
        "archived": "archived",
    }
    return mapping.get(review_status, "review_in_progress")


def hints_for(source_cost_type: str) -> dict[str, str]:
    if source_cost_type == "purchase_cny":
        return {
            "exchangeRate": "최근 적용 환율",
            "purchaseCny": "확정 발주 단가",
        }
    if source_cost_type == "purchase_krw":
        return {
            "purchaseKrw": "확정 매입 단가",
        }
    if source_cost_type == "cogs":
        return {
            "cogs": "확정 단위 원가",
        }
    return {}


def generate_product_id(product_name: str, branduid: str, sheet_name: str, row_number: int) -> str:
    if branduid:
        return f"pressco21-{branduid}"
    safe_name = "".join(char for char in product_name.lower() if char.isalnum())[:24] or f"{sheet_name}-{row_number}"
    return f"pressco21-manual-{safe_name}"


def load_base_profiles(workbook) -> dict[str, Any]:
    ws = workbook["base_profiles"]
    profiles = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        base_profile_id = normalize_text(row[0])
        if not base_profile_id:
            continue
        profiles[base_profile_id] = {
            "label": normalize_text(row[1]),
            "procurementType": normalize_text(row[2]),
            "sourceCostType": normalize_text(row[3]),
            "fixedDefaults": {
                "branchManagementRate": parse_number(row[4]) or 0,
                "freightPerUnit": parse_number(row[5]) or 0,
                "tariffRate": parse_number(row[6]) or 0,
                "packagingCost": parse_number(row[7]) or 0,
                "inspectionCost": parse_number(row[8]) or 0,
                "lossRate": parse_number(row[9]) or 0,
                "vatRate": parse_number(row[10]) or 0,
                "includeVatInCogs": parse_flag(row[11]),
                "shippingPerUnit": parse_number(row[12]) or 0,
            },
        }
    return profiles


def build_product(
    row: tuple[Any, ...],
    index: dict[str, int],
    sheet_name: str,
    row_number: int,
    base_profiles: dict[str, Any],
    seed_by_id: dict[str, Any],
    seed_by_identity: dict[tuple[str, str], dict[str, Any]],
) -> dict[str, Any] | None:
    export_flag = parse_flag(row_value(row, index, "export_flag"))
    if not export_flag:
        return None

    product_name = normalize_text(row_value(row, index, "product_name"))
    sku_code = normalize_text(row_value(row, index, "sku_code"))
    branduid = normalize_text(row_value(row, index, "branduid"))
    if not any([product_name, sku_code, branduid]):
        return None

    base_profile_id = normalize_text(row_value(row, index, "base_profile_id"))
    base_profile = base_profiles.get(base_profile_id)
    if not base_profile:
        raise ValueError(f"{sheet_name}!{row_number}: base_profile_id '{base_profile_id}' not found")

    product_id = normalize_text(row_value(row, index, "product_id")) or generate_product_id(product_name, branduid, sheet_name, row_number)
    source_cost_type = normalize_text(row_value(row, index, "source_cost_type")) or base_profile.get("sourceCostType", "")
    review_status = normalize_text(row_value(row, index, "review_status"))
    status = status_from_review(review_status)
    seed_product = seed_by_id.get(product_id) or seed_by_identity.get((branduid, product_name)) or {}
    snapshot = dict(seed_product.get("latestSourceSnapshot", {}))

    source_cost = parse_number(row_value(row, index, "source_cost_input"))
    if source_cost is None:
        source_cost = parse_number(row_value(row, index, "source_cost_seed"))
    selling_price = parse_number(row_value(row, index, "selling_price_input"))
    if selling_price is None:
        selling_price = parse_number(row_value(row, index, "selling_price_seed"))

    if source_cost_type == "purchase_cny" and source_cost is not None:
        snapshot["latestPurchaseCny"] = source_cost
        snapshot["costHistoryCny"] = parse_history(row_value(row, index, "seed_cost_history")) or snapshot.get("costHistoryCny", [])
    elif source_cost_type == "purchase_krw" and source_cost is not None:
        snapshot["latestPurchaseKrw"] = source_cost
    elif source_cost_type == "cogs" and source_cost is not None:
        snapshot["latestCogsKrw"] = source_cost

    if selling_price is not None:
        snapshot["latestSellingPriceKrw"] = selling_price
    if "latestOrderDate" not in snapshot:
        snapshot["latestOrderDate"] = normalize_text(row_value(row, index, "verified_at"))

    fixed_overrides = {}
    for field_name, column_key in FIELD_COLUMNS.items():
        row_value_numeric = row_value(row, index, column_key)
        if field_name == "includeVatInCogs":
            current = parse_flag(row_value_numeric)
            default = bool(base_profile.get("fixedDefaults", {}).get(field_name, False))
        else:
            current = parse_number(row_value_numeric)
            default = parse_number(base_profile.get("fixedDefaults", {}).get(field_name)) or 0
        if current is None:
            continue
        if current != default:
            fixed_overrides[field_name] = current

    notes = []
    seed_notes = seed_product.get("notes", [])
    notes.extend(seed_notes)
    review_note = normalize_text(row_value(row, index, "review_note"))
    if review_note and review_note != " | ".join(seed_notes):
        notes.append(review_note)
    notes.append(f"원가원장반영:{sheet_name}!{row_number}")

    aliases = unique(
        (seed_product.get("aliases") or [])
        + split_aliases(row_value(row, index, "alias_summary"))
    )
    sku_codes = unique([sku_code] + (seed_product.get("skuCodes") or []))
    branduids = unique([branduid] + (seed_product.get("branduids") or []))
    sabang_codes = unique([normalize_text(row_value(row, index, "sabang_code"))] + (seed_product.get("sabangCodes") or []))

    product = {
        "productId": product_id,
        "status": status,
        "productName": product_name or seed_product.get("productName", ""),
        "aliases": aliases,
        "skuCodes": sku_codes,
        "branduids": branduids,
        "sabangCodes": sabang_codes,
        "category": normalize_text(row_value(row, index, "category")) or seed_product.get("category", ""),
        "baseProfileId": base_profile_id,
        "sourceCostType": source_cost_type,
        "latestSourceSnapshot": snapshot,
        "fixedOverrides": fixed_overrides,
        "variableInputHints": hints_for(source_cost_type),
        "verifiedAt": normalize_text(row_value(row, index, "verified_at")),
        "verifiedBy": normalize_text(row_value(row, index, "verified_by")),
        "sourceConfidence": normalize_text(row_value(row, index, "source_confidence")),
        "notes": unique(notes),
    }
    return product


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser()
    output_path = Path(args.output).expanduser()
    seed_path = Path(args.seed).expanduser() if args.seed else output_path.with_name("openclaw-pressco21-product-cost-profiles.seed.json")

    workbook = load_workbook(input_path, data_only=False)
    base_profiles = load_base_profiles(workbook)
    seed_payload, seed_by_id, seed_by_identity = load_seed(seed_path)

    products = []
    for sheet_name in ["existing_products", "new_products"]:
        ws = workbook[sheet_name]
        index = build_sheet_index(ws)
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            product = build_product(row, index, sheet_name, row_number, base_profiles, seed_by_id, seed_by_identity)
            if product is not None:
                products.append(product)

    payload = {
        "version": "0.2",
        "updatedAt": date.today().isoformat(),
        "description": "PRESSCO21 reviewed cost profile master exported from the spreadsheet ledger.",
        "sourceWorkbook": str(input_path),
        "selectionRule": seed_payload.get("selectionRule", "reviewed spreadsheet export"),
        "baseProfiles": base_profiles,
        "products": products,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(products)} reviewed cost profiles to {output_path}")


if __name__ == "__main__":
    main()
