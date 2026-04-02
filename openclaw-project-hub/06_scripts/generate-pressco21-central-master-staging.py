#!/usr/bin/env python3

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SOURCE_HEADERS = [
    "1688 한글품명",
    "1688 규격",
    "최근단가(CNY)",
    "최근주문일",
    "주문횟수",
    "1688 URL",
    "DB코드",
    "SKU코드",
    "SKU상품명",
    "메이크샵 branduid",
    "메이크샵 상품명",
    "메이크샵 판매가",
    "메이크샵 카테고리",
    "매핑상태",
    "사방넷 상품코드",
]

STAGING_HEADERS = [
    "canonical_product_id",
    "variant_id",
    "procurement_type",
    "base_profile_id",
    "tariff_rule_class",
    "tariff_rate_override",
    "origin_certificate_default",
    "verified_ready",
    "last_verified_cogs",
    "last_verified_at",
    "last_verified_by",
    "review_note",
]

HEADER_FILL_SOURCE = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL_STAGING = PatternFill("solid", fgColor="FCE4D6")
HEADER_FILL_GUIDE = PatternFill("solid", fgColor="E2F0D9")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a staging central-product-master workbook from SKU master.")
    parser.add_argument("--input", required=True, help="Source SKU master workbook")
    parser.add_argument("--output", required=True, help="Output staging workbook path")
    return parser.parse_args()


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def infer_procurement_type() -> str:
    return "china_import"


def infer_base_profile(category: str, product_name: str, sku_name: str) -> str:
    text = " ".join([category, product_name, sku_name]).lower()
    if "dry leaf" in text or "드라이 리프" in text:
        return "china_import_dry_leaf_special"
    if "dry cut flower" in text or "드라이 컷 플라워" in text:
        return "china_import_dry_cut_flower_special"
    if any(token in text for token in ["레진", "resin", "silicone", "실리콘", "용액", "액체"]):
        return "china_import_resin_liquid"
    if any(token in text for token in ["유리", "glass", "orb", "볼", "구슬"]):
        return "china_import_fragile_glass"
    if any(token in text for token in ["몰드", "mold", "도구", "tool", "자석", "magnet", "케이블", "cable", "music box"]):
        return "china_import_mold_tool"
    if any(token in text for token in ["압화", "보존화", "플라워", "flower", "잎", "leaf"]):
        return "china_import_floral"
    return "china_import_floral"


def infer_tariff_rule(base_profile_id: str) -> str:
    mapping = {
        "china_import_dry_leaf_special": "special_dry_leaf_3_6",
        "china_import_dry_cut_flower_special": "special_dry_cut_flower_25",
        "china_import_floral": "default_low_tariff_2_1",
        "china_import_resin_liquid": "default_zero_tariff",
        "china_import_mold_tool": "default_low_tariff_2_1",
        "china_import_fragile_glass": "default_zero_tariff",
    }
    return mapping.get(base_profile_id, "default_low_tariff_2_1")


def infer_origin_certificate_default(base_profile_id: str) -> str:
    if base_profile_id in {"china_import_dry_leaf_special", "china_import_dry_cut_flower_special"}:
        return "CHECK"
    return "Y"


def build_canonical_product_id(branduid: str, sku_code: str, row_index: int) -> str:
    if branduid:
        return f"P-{branduid}"
    if sku_code:
        return f"P-{sku_code}"
    return f"P-ROW-{row_index}"


def build_variant_id(sku_code: str, row_index: int) -> str:
    if sku_code:
        return f"V-{sku_code}"
    return f"V-ROW-{row_index}"


def style_headers(ws, source_header_count: int) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        if cell.column <= source_header_count:
            cell.fill = HEADER_FILL_SOURCE
        else:
            cell.fill = HEADER_FILL_STAGING
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def apply_widths(ws) -> None:
    widths = {
        "A": 24,
        "B": 24,
        "C": 14,
        "D": 14,
        "E": 10,
        "F": 32,
        "G": 12,
        "H": 16,
        "I": 30,
        "J": 16,
        "K": 30,
        "L": 14,
        "M": 20,
        "N": 14,
        "O": 16,
        "P": 20,
        "Q": 20,
        "R": 18,
        "S": 22,
        "T": 18,
        "U": 16,
        "V": 18,
        "W": 16,
        "X": 12,
        "Y": 18,
        "Z": 18,
        "AA": 12,
        "AB": 14,
        "AC": 14,
        "AD": 14,
        "AE": 18,
        "AF": 40,
        "AG": 16,
        "AH": 16,
        "AI": 16,
        "AJ": 18,
        "AK": 18,
        "AL": 18,
        "AM": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def add_validations(ws) -> None:
    max_row = ws.max_row
    validations = [
        ("Q2:Q{}".format(max_row), '"china_import"'),
        (
            "R2:R{}".format(max_row),
            '"china_import_floral,china_import_dry_leaf_special,china_import_dry_cut_flower_special,china_import_resin_liquid,china_import_mold_tool,china_import_fragile_glass"',
        ),
        (
            "S2:S{}".format(max_row),
            '"default_low_tariff_2_1,default_zero_tariff,special_dry_leaf_3_6,special_dry_cut_flower_25,manual_override"',
        ),
        ("U2:U{}".format(max_row), '"Y,N,CHECK"'),
        ("V2:V{}".format(max_row), '"N,Y"'),
    ]
    for cell_range, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(cell_range)


def create_guide_sheet(wb) -> None:
    if "운영가이드" in wb.sheetnames:
        del wb["운영가이드"]
    ws = wb.create_sheet("운영가이드")
    rows = [
        ["필드", "설명", "권장값"],
        ["canonical_product_id", "대표 상품 식별자", "branduid 우선, 없으면 SKU 기반"],
        ["variant_id", "옵션/변형 식별자", "SKU 기반"],
        ["procurement_type", "조달유형", "china_import"],
        ["base_profile_id", "원가 프로필", "일반 플라워/레진/몰드/유리/특별예외"],
        ["tariff_rule_class", "관세 규칙", "default_low_tariff_2_1 / special_dry_leaf_3_6 / special_dry_cut_flower_25"],
        ["tariff_rate_override", "수동 관세율", "기본은 비워두고 예외만 입력"],
        ["origin_certificate_default", "원산지증명 기본값", "Y 또는 CHECK"],
        ["verified_ready", "원가 검증 완료 여부", "Y 되기 전까지 N 유지"],
        ["last_verified_cogs", "확정 COGS", "OpenClaw 검토 후 입력"],
        ["last_verified_at", "마지막 검증일", "YYYY-MM-DD"],
        ["last_verified_by", "검증자", "이름 입력"],
        ["review_note", "예외 메모", "관세/원가 충돌 이유 기록"],
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_GUIDE
        cell.alignment = CENTER
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 36


def main() -> None:
    args = parse_args()
    source_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(source_path)
    ws = wb[wb.sheetnames[0]]

    headers = [normalize_text(cell.value) for cell in ws[1]]
    if headers[: len(SOURCE_HEADERS)] != SOURCE_HEADERS:
        raise ValueError("Unexpected source workbook headers")

    source_header_count = len(headers)
    for header in STAGING_HEADERS:
        ws.cell(row=1, column=ws.max_column + 1, value=header)

    column_index = {normalize_text(cell.value): cell.column for cell in ws[1]}
    for row_index in range(2, ws.max_row + 1):
        branduid = normalize_text(ws.cell(row=row_index, column=column_index["메이크샵 branduid"]).value)
        sku_code = normalize_text(ws.cell(row=row_index, column=column_index["SKU코드"]).value)
        product_name = normalize_text(ws.cell(row=row_index, column=column_index["메이크샵 상품명"]).value)
        sku_name = normalize_text(ws.cell(row=row_index, column=column_index["SKU상품명"]).value)
        category = normalize_text(ws.cell(row=row_index, column=column_index["메이크샵 카테고리"]).value)

        base_profile_id = infer_base_profile(category, product_name, sku_name)
        values = {
            "canonical_product_id": build_canonical_product_id(branduid, sku_code, row_index),
            "variant_id": build_variant_id(sku_code, row_index),
            "procurement_type": infer_procurement_type(),
            "base_profile_id": base_profile_id,
            "tariff_rule_class": infer_tariff_rule(base_profile_id),
            "tariff_rate_override": "",
            "origin_certificate_default": infer_origin_certificate_default(base_profile_id),
            "verified_ready": "N",
            "last_verified_cogs": "",
            "last_verified_at": "",
            "last_verified_by": "",
            "review_note": "",
        }
        for header, value in values.items():
            ws.cell(row=row_index, column=column_index[header], value=value)

    style_headers(ws, source_header_count)
    apply_widths(ws)
    add_validations(ws)
    create_guide_sheet(wb)
    wb.save(output_path)


if __name__ == "__main__":
    main()
