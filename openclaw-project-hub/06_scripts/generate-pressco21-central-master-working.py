#!/usr/bin/env python3

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SOURCE_HEADERS = [
    "1688 한글품명",
    "1688 규격",
    "최근단가(CNY)",
    "최근주문일",
    "주문횟수",
    "1688 URL",
    "DB코드",
    "SKU코드",
    "SKU상품명",
    "메이크샵 branduid",
    "메이크샵 상품명",
    "메이크샵 판매가",
    "메이크샵 카테고리",
    "매핑상태",
    "사방넷 상품코드",
]

HEADERS = [
    "대표상품ID",
    "옵션ID",
    "상품명",
    "옵션/SKU상품명",
    "SKU코드",
    "메이크샵 branduid",
    "사방넷 상품코드",
    "카테고리",
    "최근단가(CNY)",
    "메이크샵 판매가",
    "기본프로필",
    "관세규칙",
    "검증완료",
    "메모",
    "원가타입",
    "확정원가입력",
    "적용환율",
    "배분수량",
    "개당CBM",
    "원산지증명",
    "관세율수동",
    "국내입고배송비수동",
    "포장비수동",
    "검품비수동",
    "손실충당율수동",
    "판매가입력",
    "채널",
    "채널수수료수동",
    "광고비율수동",
    "추천관세율",
    "추천국내입고배송비",
    "추천포장비",
    "추천검품비",
    "추천손실충당율",
    "추천지사관리비율",
    "추천원산지증명비(건)",
    "추천채널수수료",
    "추천광고비율",
    "적용관세율",
    "적용국내입고배송비",
    "적용포장비",
    "적용검품비",
    "적용손실충당율",
    "적용채널수수료",
    "적용광고비율",
    "매입원화",
    "지사관리비",
    "해상운임",
    "창고작업비",
    "통관수수료배분",
    "원산지증명비배분",
    "운임합계",
    "관세",
    "손실충당금",
    "예상COGS",
    "공급가",
    "채널수수료",
    "광고비",
    "예상마진액",
    "예상원가율",
    "예상마진율",
    "판정",
]

HEADER_FILL_INFO = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL_LOOKUP = PatternFill("solid", fgColor="EDEDED")
HEADER_FILL_OUTPUT = PatternFill("solid", fgColor="E2F0D9")
HEADER_FILL_GUIDE = PatternFill("solid", fgColor="DDEBF7")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")

PROFILE_ROWS = [
    ("china_import_floral", "중국 사입 - 플라워/프리저브드 일반형", 0.021, 150, 150, 50, 0.03, 0.10, 82000, 6000, 33000, 50000),
    ("china_import_dry_leaf_special", "중국 사입 - 드라이 리프 특별형", 0.036, 150, 150, 50, 0.03, 0.10, 82000, 6000, 33000, 50000),
    ("china_import_dry_cut_flower_special", "중국 사입 - 드라이 컷 플라워 특별형", 0.25, 150, 150, 50, 0.03, 0.10, 82000, 6000, 33000, 50000),
    ("china_import_resin_liquid", "중국 사입 - 레진/용액 기본형", 0.0, 180, 200, 80, 0.04, 0.10, 82000, 6000, 33000, 50000),
    ("china_import_mold_tool", "중국 사입 - 몰드/도구 기본형", 0.021, 120, 120, 50, 0.02, 0.10, 82000, 6000, 33000, 50000),
    ("china_import_fragile_glass", "중국 사입 - 유리/깨짐주의 기본형", 0.0, 220, 250, 120, 0.05, 0.10, 82000, 6000, 33000, 50000),
]

CHANNEL_ROWS = [
    ("메이크샵 자사몰", 0.0385, "", "광고비율은 아직 미정"),
    ("네이버 스마트스토어", 0.0574, "", "광고비율은 아직 미정"),
    ("쿠팡 윙", 0.1138, 0.15, "광고비 15% 고정"),
    ("11번가", 0.12, 0.0, "광고 거의 없음"),
    ("오늘의집", 0.12, "", "추후 별도 검토"),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a non-technical working workbook for PRESSCO21 central product master.")
    parser.add_argument("--input", required=True, help="Source SKU master workbook path")
    parser.add_argument("--output", required=True, help="Output workbook path")
    return parser.parse_args()


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def infer_base_profile(category: str, product_name: str, sku_name: str) -> str:
    text = " ".join([category, product_name, sku_name]).lower()
    if "dry leaf" in text or "드라이 리프" in text:
        return "china_import_dry_leaf_special"
    if "dry cut flower" in text or "드라이 컷 플라워" in text:
        return "china_import_dry_cut_flower_special"
    if any(token in text for token in ["레진", "resin", "silicone", "실리콘", "용액", "액체"]):
        return "china_import_resin_liquid"
    if any(token in text for token in ["유리", "glass", "orb", "볼", "구슬"]):
        return "china_import_fragile_glass"
    if any(token in text for token in ["몰드", "mold", "도구", "tool", "자석", "magnet", "케이블", "cable", "music box"]):
        return "china_import_mold_tool"
    return "china_import_floral"


def infer_tariff_rule(base_profile_id: str) -> str:
    mapping = {
        "china_import_dry_leaf_special": "special_dry_leaf_3_6",
        "china_import_dry_cut_flower_special": "special_dry_cut_flower_25",
        "china_import_floral": "default_low_tariff_2_1",
        "china_import_resin_liquid": "default_zero_tariff",
        "china_import_mold_tool": "default_low_tariff_2_1",
        "china_import_fragile_glass": "default_zero_tariff",
    }
    return mapping.get(base_profile_id, "default_low_tariff_2_1")


def infer_origin_certificate(base_profile_id: str) -> str:
    if base_profile_id in {"china_import_dry_leaf_special", "china_import_dry_cut_flower_special"}:
        return "CHECK"
    return "Y"


def build_product_id(branduid: str, sku_code: str, row_num: int) -> str:
    if branduid:
        return f"P-{branduid}"
    if sku_code:
        return f"P-{sku_code}"
    return f"P-ROW-{row_num}"


def build_variant_id(sku_code: str, row_num: int) -> str:
    if sku_code:
        return f"V-{sku_code}"
    return f"V-ROW-{row_num}"


def style_product_sheet(ws) -> None:
    info_cols = range(1, 15)
    input_cols = range(15, 30)
    lookup_cols = range(30, 39)
    output_cols = range(39, 63)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        if cell.column in info_cols:
            cell.fill = HEADER_FILL_INFO
        elif cell.column in input_cols:
            cell.fill = HEADER_FILL_INPUT
        elif cell.column in lookup_cols:
            cell.fill = HEADER_FILL_LOOKUP
        elif cell.column in output_cols:
            cell.fill = HEADER_FILL_OUTPUT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 18, "B": 18, "C": 28, "D": 28, "E": 16, "F": 16, "G": 16, "H": 18, "I": 14, "J": 14,
        "K": 22, "L": 20, "M": 12, "N": 24, "O": 12, "P": 14, "Q": 12, "R": 10, "S": 10, "T": 12,
        "U": 12, "V": 16, "W": 12, "X": 12, "Y": 12, "Z": 14, "AA": 16, "AB": 14, "AC": 12, "AD": 12,
        "AE": 16, "AF": 12, "AG": 12, "AH": 12, "AI": 12, "AJ": 16, "AK": 14, "AL": 12, "AM": 12,
        "AN": 16, "AO": 12, "AP": 12, "AQ": 12, "AR": 14, "AS": 12, "AT": 12, "AU": 12, "AV": 12,
        "AW": 12, "AX": 14, "AY": 14, "AZ": 12, "BA": 12, "BB": 12, "BC": 14, "BD": 12, "BE": 12,
        "BF": 12, "BG": 12, "BH": 12, "BI": 12, "BJ": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_validations(ws) -> None:
    max_row = ws.max_row
    validations = [
        ("K2:K{}".format(max_row), '"china_import_floral,china_import_dry_leaf_special,china_import_dry_cut_flower_special,china_import_resin_liquid,china_import_mold_tool,china_import_fragile_glass"'),
        ("M2:M{}".format(max_row), '"N,Y"'),
        ("O2:O{}".format(max_row), '"purchase_cny,purchase_krw,cogs"'),
        ("T2:T{}".format(max_row), '"Y,N,CHECK"'),
        ("AA2:AA{}".format(max_row), '"메이크샵 자사몰,네이버 스마트스토어,쿠팡 윙,11번가,오늘의집"'),
    ]
    for cell_range, formula in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(cell_range)


def set_formula(ws, row: int, col: int, formula: str) -> None:
    ws.cell(row=row, column=col, value=formula)


def create_profile_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("프로필기준")
    ws.append([
        "base_profile_id",
        "설명",
        "추천관세율",
        "추천국내입고배송비",
        "추천포장비",
        "추천검품비",
        "추천손실충당율",
        "추천지사관리비율",
        "해상운임(CBM)",
        "창고작업비(CBM)",
        "통관수수료(건)",
        "원산지증명비(건)",
    ])
    for row in PROFILE_ROWS:
        ws.append(row)
    start_col = 14
    for offset, header in enumerate(["채널", "추천채널수수료", "추천광고비율", "메모"]):
        ws.cell(row=1, column=start_col + offset, value=header)
    for idx, row in enumerate(CHANNEL_ROWS, start=2):
        for offset, value in enumerate(row):
            ws.cell(row=idx, column=start_col + offset, value=value)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL_GUIDE
    ws.freeze_panes = "A2"
    for col in range(1, 18):
        ws.column_dimensions[get_column_letter(col)].width = 18


def create_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("입력설명")
    rows = [
        ["단계", "직원이 하는 일", "설명"],
        ["1", "상품 찾기", "상품명, SKU, 메이크샵 branduid를 확인한다."],
        ["2", "노란칸만 수정", "확정원가입력, 배분수량, 개당CBM, 판매가입력만 우선 채운다."],
        ["3", "예외만 수동 입력", "관세율/포장비/검품비/손실충당율은 특별한 경우만 수동 입력한다."],
        ["4", "초록칸 확인", "예상COGS, 예상원가율, 예상마진율, 판정을 본다."],
        ["5", "검증완료 표시", "실제 검토가 끝난 상품만 검증완료를 Y로 바꾼다."],
        ["핵심", "노란칸 = 직접 입력", "회색칸은 시스템 추천값, 초록칸은 자동 계산 결과다."],
        ["주의", "드라이 컷 플라워", "꽃류는 특별예외 관세 25% 여부를 반드시 확인한다."],
        ["주의", "드라이 리프", "잎류는 현재 2025 서류 기준 특별예외 3.6%를 사용한다."],
        ["주의", "자사몰/스마트스토어 광고비", "아직 고정값이 아니므로 필요할 때만 직접 입력한다."],
        ["주의", "쿠팡 광고비", "쿠팡 윙은 15%가 기본 추천값으로 들어간다."],
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL_GUIDE
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 56


def create_dashboard_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("요약대시보드")
    rows = [
        ("지표", "값", "설명"),
        ("총 상품 수", "=COUNTA(상품마스터_작업!A:A)-1", "전체 상품 행 수"),
        ("검증완료 수", '=COUNTIF(상품마스터_작업!M:M,"Y")', "검증 완료된 상품 수"),
        ("검토대기 수", '=COUNTIF(상품마스터_작업!M:M,"N")', "아직 검토중인 상품 수"),
        ("특별예외 관세 수", '=COUNTIF(상품마스터_작업!L:L,"special*")', "드라이 리프/꽃 등 예외군"),
        ("예상COGS 계산완료 수", '=COUNTIF(상품마스터_작업!BC:BC,">0")', "실제 계산 결과가 나온 상품 수"),
        ("평균 예상원가율", '=IFERROR(AVERAGEIF(상품마스터_작업!BH:BH,">0"),"")', "판매가 대비 COGS 비율"),
        ("평균 예상마진율", '=IFERROR(AVERAGEIF(상품마스터_작업!BI:BI,">0"),"")', "채널비용/광고비 반영 후 비율"),
        ("위험 상품 수", '=COUNTIF(상품마스터_작업!BJ:BJ,"위험")', "즉시 재검토 필요"),
        ("검토필요 상품 수", '=COUNTIF(상품마스터_작업!BJ:BJ,"검토필요")', "보완 검토 필요"),
    ]
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL_GUIDE
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 36
    for row in [7, 8]:
        ws.cell(row=row, column=2).number_format = "0.0%"


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    source_wb = load_workbook(input_path, read_only=True, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[0]]
    source_headers = [normalize_text(c) for c in next(source_ws.iter_rows(values_only=True))]
    if source_headers[: len(SOURCE_HEADERS)] != SOURCE_HEADERS:
        raise ValueError("Unexpected source workbook headers")

    wb = Workbook()
    ws = wb.active
    ws.title = "상품마스터_작업"
    ws.append(HEADERS)

    for row_num, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
        branduid = normalize_text(row[9])
        sku_code = normalize_text(row[7])
        sku_name = normalize_text(row[8])
        product_name = normalize_text(row[10]) or sku_name or normalize_text(row[0])
        category = normalize_text(row[12])
        selling_price = row[11]
        recent_cost = row[2]
        base_profile_id = infer_base_profile(category, product_name, sku_name)

        ws.append([
            build_product_id(branduid, sku_code, row_num),
            build_variant_id(sku_code, row_num),
            product_name,
            sku_name,
            sku_code,
            branduid,
            normalize_text(row[14]),
            category,
            recent_cost,
            selling_price,
            base_profile_id,
            infer_tariff_rule(base_profile_id),
            "N",
            "",
            "purchase_cny",
            recent_cost,
            218.340611,
            "",
            "",
            infer_origin_certificate(base_profile_id),
            "",
            "",
            "",
            "",
            "",
            selling_price,
            "메이크샵 자사몰",
            "",
            "",
        ])

        current_row = ws.max_row
        set_formula(ws, current_row, 30, '=IFERROR(VLOOKUP($K{0},프로필기준!$A$2:$L$20,3,FALSE),"")'.format(current_row))
        set_formula(ws, current_row, 31, '=IFERROR(VLOOKUP($K{0},프로필기준!$A$2:$L$20,4,FALSE),"")'.format(current_row))
        set_formula(ws, current_row, 32, '=IFERROR(VLOOKUP($K{0},프로필기준!$A$2:$L$20,5,FALSE),"")'.format(current_row))
        set_formula(ws, current_row, 33, '=IFERROR(VLOOKUP($K{0},프로필기준!$A$2:$L$20,6,FALSE),"")'.format(current_row))
        set_formula(ws, current_row, 34, '=IFERROR(VLOOKUP($K{0},프로필기준!$A$2:$L$20,7,FALSE),"")'.format(current_row))
        set_formula(ws, current_row, 35, '=IFERROR(VLOOKUP($K{0},프로필기준!$A$2:$L$20,8,FALSE),"")'.format(current_row))
        set_formula(ws, current_row, 36, '=IFERROR(VLOOKUP($K{0},프로필기준!$A$2:$L$20,12,FALSE),"")'.format(current_row))
        set_formula(ws, current_row, 37, '=IFERROR(VLOOKUP($AA{0},프로필기준!$N$2:$Q$10,2,FALSE),"")'.format(current_row))
        set_formula(ws, current_row, 38, '=IFERROR(VLOOKUP($AA{0},프로필기준!$N$2:$Q$10,3,FALSE),"")'.format(current_row))

        set_formula(ws, current_row, 39, '=IF($U{0}<>"",$U{0},$AD{0})'.format(current_row))
        set_formula(ws, current_row, 40, '=IF($V{0}<>"",$V{0},$AE{0})'.format(current_row))
        set_formula(ws, current_row, 41, '=IF($W{0}<>"",$W{0},$AF{0})'.format(current_row))
        set_formula(ws, current_row, 42, '=IF($X{0}<>"",$X{0},$AG{0})'.format(current_row))
        set_formula(ws, current_row, 43, '=IF($Y{0}<>"",$Y{0},$AH{0})'.format(current_row))
        set_formula(ws, current_row, 44, '=IF($AB{0}<>"",$AB{0},$AK{0})'.format(current_row))
        set_formula(ws, current_row, 45, '=IF($AC{0}<>"",$AC{0},$AL{0})'.format(current_row))

        set_formula(ws, current_row, 46, '=IF($O{0}="purchase_cny",ROUND($P{0}*$Q{0},0),IF($O{0}="purchase_krw",$P{0},IF($O{0}="cogs",$P{0},"")))'.format(current_row))
        set_formula(ws, current_row, 47, '=IF($AT{0}="","",ROUND($AT{0}*$AI{0},0))'.format(current_row))
        set_formula(ws, current_row, 48, '=IF(AND($S{0}<>"",$S{0}>0),ROUND($S{0}*VLOOKUP($K{0},프로필기준!$A$2:$L$20,9,FALSE),0),0)'.format(current_row))
        set_formula(ws, current_row, 49, '=IF(AND($S{0}<>"",$S{0}>0),ROUND($S{0}*VLOOKUP($K{0},프로필기준!$A$2:$L$20,10,FALSE),0),0)'.format(current_row))
        set_formula(ws, current_row, 50, '=IF(AND($R{0}<>"",$R{0}>0),ROUND(VLOOKUP($K{0},프로필기준!$A$2:$L$20,11,FALSE)/$R{0},0),0)'.format(current_row))
        set_formula(ws, current_row, 51, '=IF(AND($R{0}<>"",$R{0}>0,$T{0}="Y"),ROUND($AJ{0}/$R{0},0),0)'.format(current_row))
        set_formula(ws, current_row, 52, '=SUM($AN{0},$AV{0},$AW{0},$AX{0},$AY{0})'.format(current_row))
        set_formula(ws, current_row, 53, '=IF($AT{0}="","",ROUND(($AT{0}+$AV{0}+$AW{0})*$AM{0},0))'.format(current_row))
        set_formula(ws, current_row, 54, '=IF($AT{0}="","",ROUND(($AT{0}+$AU{0}+$AZ{0}+$BA{0}+$AO{0}+$AP{0})*$AQ{0},0))'.format(current_row))
        set_formula(ws, current_row, 55, '=IF($O{0}="cogs",$P{0},SUM($AT{0},$AU{0},$AZ{0},$BA{0},$AO{0},$AP{0},$BB{0}))'.format(current_row))
        set_formula(ws, current_row, 56, '=IF($Z{0}="","",ROUND($Z{0}/1.1,0))'.format(current_row))
        set_formula(ws, current_row, 57, '=IF($BD{0}="","",ROUND($BD{0}*$AR{0},0))'.format(current_row))
        set_formula(ws, current_row, 58, '=IF(OR($BD{0}="",$AS{0}=""),0,ROUND($BD{0}*$AS{0},0))'.format(current_row))
        set_formula(ws, current_row, 59, '=IF($BD{0}="","",$BD{0}-$BC{0}-$BE{0}-$BF{0})'.format(current_row))
        set_formula(ws, current_row, 60, '=IF($Z{0}="","",ROUND($BC{0}/$Z{0},3))'.format(current_row))
        set_formula(ws, current_row, 61, '=IF($BD{0}="","",ROUND($BG{0}/$BD{0},3))'.format(current_row))
        set_formula(ws, current_row, 62, '=IF($BI{0}="","",IF($BI{0}>=0.3,"좋음",IF($BI{0}>=0.2,"운영가능",IF($BI{0}>=0.1,"검토필요","위험"))))'.format(current_row))

    style_product_sheet(ws)
    add_validations(ws)
    create_profile_sheet(wb)
    create_guide_sheet(wb)
    create_dashboard_sheet(wb)

    if hasattr(wb, "calculation"):
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True

    for row in range(2, ws.max_row + 1):
        for col in [17, 21, 25, 28, 29, 30, 34, 35, 37, 38, 39, 43, 44, 45, 60, 61]:
            ws.cell(row=row, column=col).number_format = "0.0%"
        for col in [16, 18, 19, 22, 23, 24, 26, 31, 32, 33, 36, 40, 41, 42, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59]:
            ws.cell(row=row, column=col).number_format = '#,##0'

    wb.save(output_path)


if __name__ == "__main__":
    main()
