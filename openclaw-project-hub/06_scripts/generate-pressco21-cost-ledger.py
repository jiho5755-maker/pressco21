#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    ("review_status", "검토상태"),
    ("export_flag", "내보내기"),
    ("product_id", "productId"),
    ("product_name", "상품명"),
    ("sku_code", "SKU"),
    ("branduid", "branduid"),
    ("sabang_code", "사방넷코드"),
    ("category", "카테고리"),
    ("base_profile_id", "기본프로필"),
    ("procurement_type", "조달유형"),
    ("source_cost_type", "원가타입"),
    ("source_cost_seed", "시드원가"),
    ("source_cost_input", "확정원가입력"),
    ("exchange_rate", "환율"),
    ("selling_price_seed", "시드판매가"),
    ("selling_price_input", "현재판매가입력"),
    ("branch_management_rate", "지사관리비율"),
    ("freight_per_unit", "단위당운임"),
    ("tariff_rate", "관세율"),
    ("packaging_cost", "포장비"),
    ("inspection_cost", "검품비"),
    ("loss_rate", "손실충당율"),
    ("vat_rate", "VAT율"),
    ("include_vat_in_cogs", "VAT포함"),
    ("shipping_per_unit", "단위당배송비"),
    ("normalized_cogs", "계산COGS"),
    ("cost_rate", "원가율"),
    ("gross_margin_won", "매출총이익"),
    ("gross_margin_rate", "매출총이익률"),
    ("source_confidence", "원가신뢰도"),
    ("verified_by", "검증자"),
    ("verified_at", "검증일"),
    ("review_note", "검토메모"),
    ("seed_cost_history", "시드원가이력"),
    ("alias_summary", "별칭/원본명"),
]

HEADER_INDEX = {key: position + 1 for position, (key, _label) in enumerate(HEADERS)}
FIELD_TO_COLUMN = {
    "branchManagementRate": "branch_management_rate",
    "freightPerUnit": "freight_per_unit",
    "tariffRate": "tariff_rate",
    "packagingCost": "packaging_cost",
    "inspectionCost": "inspection_cost",
    "lossRate": "loss_rate",
    "vatRate": "vat_rate",
    "includeVatInCogs": "include_vat_in_cogs",
    "shippingPerUnit": "shipping_per_unit",
}
DEFAULT_WIDTHS = {
    "A": 22,
    "B": 10,
    "C": 18,
    "D": 34,
    "E": 16,
    "F": 14,
    "G": 14,
    "H": 20,
    "I": 24,
    "J": 16,
    "K": 16,
    "L": 12,
    "M": 12,
    "N": 10,
    "O": 12,
    "P": 12,
    "Q": 12,
    "R": 12,
    "S": 12,
    "T": 12,
    "U": 12,
    "V": 12,
    "W": 10,
    "X": 10,
    "Y": 12,
    "Z": 12,
    "AA": 10,
    "AB": 12,
    "AC": 12,
    "AD": 16,
    "AE": 12,
    "AF": 12,
    "AG": 38,
    "AH": 24,
    "AI": 24,
}
REVIEW_STATUS_OPTIONS = [
    "seeded_review_pending",
    "conflict_needs_review",
    "review_in_progress",
    "verified_ready",
    "new_draft",
    "archived",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a human-reviewable PRESSCO21 cost ledger workbook.")
    parser.add_argument("--seed", required=True, help="Path to seed product cost profiles JSON")
    parser.add_argument("--output", required=True, help="Path to output XLSX workbook")
    parser.add_argument("--new-rows-per-profile", type=int, default=3, help="Blank template rows to create for each base profile")
    return parser.parse_args()


def parse_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    try:
        numeric = float(value)
    except Exception:
        return None
    if numeric.is_integer():
        return int(numeric)
    return round(numeric, 4)


def join_list(values: list[Any], separator: str = " | ") -> str:
    return separator.join(str(value).strip() for value in values if str(value).strip())


def bool_to_flag(value: Any) -> str:
    return "Y" if bool(value) else "N"


def review_status_from_seed(status: str) -> str:
    if status == "draft_conflict_from_master":
        return "conflict_needs_review"
    if status == "draft_seeded_from_master":
        return "seeded_review_pending"
    if status == "verified_ready":
        return "verified_ready"
    return "review_in_progress"


def rate_value(base_profile: dict[str, Any], product: dict[str, Any], field_name: str) -> Any:
    fixed_defaults = deepcopy(base_profile.get("fixedDefaults", {}))
    fixed_defaults.update(product.get("fixedOverrides", {}))
    return fixed_defaults.get(field_name)


def source_cost_seed(snapshot: dict[str, Any], source_cost_type: str) -> float | int | None:
    if source_cost_type == "purchase_cny":
        return parse_number(snapshot.get("latestPurchaseCny"))
    if source_cost_type == "purchase_krw":
        return parse_number(snapshot.get("latestPurchaseKrw"))
    if source_cost_type == "cogs":
        return parse_number(snapshot.get("latestCogsKrw"))
    return None


def build_existing_rows(seed: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    base_profiles = seed.get("baseProfiles", {})
    for product in seed.get("products", []):
        base_profile_id = product.get("baseProfileId", "")
        base_profile = base_profiles.get(base_profile_id, {})
        source_cost_type = base_profile.get("sourceCostType", "")
        snapshot = product.get("latestSourceSnapshot", {})
        seed_cost = source_cost_seed(snapshot, source_cost_type)
        rows.append(
            {
                "review_status": review_status_from_seed(product.get("status", "")),
                "export_flag": "Y",
                "product_id": product.get("productId", ""),
                "product_name": product.get("productName", ""),
                "sku_code": (product.get("skuCodes") or [""])[0],
                "branduid": (product.get("branduids") or [""])[0],
                "sabang_code": (product.get("sabangCodes") or [""])[0],
                "category": product.get("category", ""),
                "base_profile_id": base_profile_id,
                "procurement_type": base_profile.get("procurementType", ""),
                "source_cost_type": source_cost_type,
                "source_cost_seed": seed_cost,
                "source_cost_input": seed_cost if product.get("status") != "draft_conflict_from_master" else None,
                "exchange_rate": 218.340611 if source_cost_type == "purchase_cny" else None,
                "selling_price_seed": parse_number(snapshot.get("latestSellingPriceKrw")),
                "selling_price_input": parse_number(snapshot.get("latestSellingPriceKrw")),
                "branch_management_rate": rate_value(base_profile, product, "branchManagementRate"),
                "freight_per_unit": rate_value(base_profile, product, "freightPerUnit"),
                "tariff_rate": rate_value(base_profile, product, "tariffRate"),
                "packaging_cost": rate_value(base_profile, product, "packagingCost"),
                "inspection_cost": rate_value(base_profile, product, "inspectionCost"),
                "loss_rate": rate_value(base_profile, product, "lossRate"),
                "vat_rate": rate_value(base_profile, product, "vatRate"),
                "include_vat_in_cogs": bool_to_flag(rate_value(base_profile, product, "includeVatInCogs")),
                "shipping_per_unit": rate_value(base_profile, product, "shippingPerUnit"),
                "source_confidence": "stable_from_seed" if product.get("status") != "draft_conflict_from_master" else "conflict_from_seed",
                "verified_by": product.get("verifiedBy", ""),
                "verified_at": product.get("verifiedAt", ""),
                "review_note": join_list(product.get("notes", [])),
                "seed_cost_history": join_list(snapshot.get("costHistoryCny", []), ", "),
                "alias_summary": join_list(product.get("aliases", [])),
            }
        )
    return rows


def build_new_rows(seed: dict[str, Any], rows_per_profile: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for base_profile_id, base_profile in seed.get("baseProfiles", {}).items():
        fixed_defaults = base_profile.get("fixedDefaults", {})
        for index in range(rows_per_profile):
            rows.append(
                {
                    "review_status": "new_draft",
                    "export_flag": "N",
                    "product_id": "",
                    "product_name": "",
                    "sku_code": "",
                    "branduid": "",
                    "sabang_code": "",
                    "category": "",
                    "base_profile_id": base_profile_id,
                    "procurement_type": base_profile.get("procurementType", ""),
                    "source_cost_type": base_profile.get("sourceCostType", ""),
                    "source_cost_seed": None,
                    "source_cost_input": None,
                    "exchange_rate": 218.340611 if base_profile.get("sourceCostType") == "purchase_cny" else None,
                    "selling_price_seed": None,
                    "selling_price_input": None,
                    "branch_management_rate": fixed_defaults.get("branchManagementRate"),
                    "freight_per_unit": fixed_defaults.get("freightPerUnit"),
                    "tariff_rate": fixed_defaults.get("tariffRate"),
                    "packaging_cost": fixed_defaults.get("packagingCost"),
                    "inspection_cost": fixed_defaults.get("inspectionCost"),
                    "loss_rate": fixed_defaults.get("lossRate"),
                    "vat_rate": fixed_defaults.get("vatRate"),
                    "include_vat_in_cogs": bool_to_flag(fixed_defaults.get("includeVatInCogs")),
                    "shipping_per_unit": fixed_defaults.get("shippingPerUnit"),
                    "source_confidence": "new_profile_template",
                    "verified_by": "",
                    "verified_at": "",
                    "review_note": f"새 상품 입력용 템플릿 {index + 1}",
                    "seed_cost_history": "",
                    "alias_summary": base_profile.get("label", ""),
                }
            )
    return rows


def formula_for(row_index: int, key: str) -> str:
    def ref(name: str) -> str:
        return f"{get_column_letter(HEADER_INDEX[name])}{row_index}"

    if key == "normalized_cogs":
        return (
            f'=IF(OR({ref("source_cost_input")}="",AND({ref("source_cost_type")}="purchase_cny",{ref("exchange_rate")}="")),"",'
            f'IF({ref("source_cost_type")}="purchase_cny",'
            f'ROUND(((({ref("source_cost_input")}*{ref("exchange_rate")})*(1+{ref("branch_management_rate")})'
            f'+{ref("freight_per_unit")}+(({ref("source_cost_input")}*{ref("exchange_rate")})+{ref("freight_per_unit")})*{ref("tariff_rate")}'
            f'+{ref("packaging_cost")}+{ref("inspection_cost")})*(1+{ref("loss_rate")})*IF(OR({ref("include_vat_in_cogs")}="Y",{ref("include_vat_in_cogs")}=TRUE,{ref("include_vat_in_cogs")}=1),1+{ref("vat_rate")},1),0),'
            f'IF({ref("source_cost_type")}="purchase_krw",'
            f'ROUND((({ref("source_cost_input")}+{ref("shipping_per_unit")}+{ref("packaging_cost")}+{ref("inspection_cost")})*(1+{ref("loss_rate")})),0),'
            f'IF({ref("source_cost_type")}="cogs",ROUND({ref("source_cost_input")},0),""))))'
        )
    if key == "cost_rate":
        return f'=IF(OR({ref("normalized_cogs")}="",{ref("selling_price_input")}=""),"",ROUND({ref("normalized_cogs")}/{ref("selling_price_input")},4))'
    if key == "gross_margin_won":
        return f'=IF(OR({ref("normalized_cogs")}="",{ref("selling_price_input")}=""),"",ROUND({ref("selling_price_input")}-{ref("normalized_cogs")},0))'
    if key == "gross_margin_rate":
        return f'=IF(OR({ref("gross_margin_won")}="",{ref("selling_price_input")}=""),"",ROUND({ref("gross_margin_won")}/{ref("selling_price_input")},4))'
    return ""


def add_header_styles(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column_index, (_key, label) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=column_index, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def apply_sheet_layout(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    for column, width in DEFAULT_WIDTHS.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 24


def style_data_cell(cell, key: str) -> None:
    input_fill = PatternFill("solid", fgColor="DBEAFE")
    formula_fill = PatternFill("solid", fgColor="DCFCE7")
    caution_fill = PatternFill("solid", fgColor="FFEDD5")
    static_fill = PatternFill("solid", fgColor="F1F5F9")
    border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    cell.border = border
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if key in {"source_cost_input", "selling_price_input", "exchange_rate", "verified_by", "verified_at", "review_note"}:
        cell.fill = input_fill
    elif key in {"normalized_cogs", "cost_rate", "gross_margin_won", "gross_margin_rate"}:
        cell.fill = formula_fill
    elif key in {"review_status", "export_flag", "source_confidence"}:
        cell.fill = caution_fill
    else:
        cell.fill = static_fill
    if key.endswith("_rate"):
        cell.number_format = "0.0%"
    elif key in {"source_cost_seed", "source_cost_input", "exchange_rate", "selling_price_seed", "selling_price_input", "freight_per_unit", "packaging_cost", "inspection_cost", "shipping_per_unit", "normalized_cogs", "gross_margin_won"}:
        cell.number_format = "#,##0"
    if key == "verified_at":
        cell.number_format = "yyyy-mm-dd"


def add_validations(ws, max_row: int) -> None:
    status_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(REVIEW_STATUS_OPTIONS)}"',
        allow_blank=False,
    )
    export_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    vat_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(export_validation)
    ws.add_data_validation(vat_validation)
    status_validation.add(f"A2:A{max_row}")
    export_validation.add(f"B2:B{max_row}")
    vat_validation.add(f"X2:X{max_row}")


def write_rows(ws, rows: list[dict[str, Any]]) -> None:
    add_header_styles(ws)
    apply_sheet_layout(ws)
    for row_index, row_data in enumerate(rows, start=2):
        for key, _label in HEADERS:
            column_index = HEADER_INDEX[key]
            cell = ws.cell(row=row_index, column=column_index)
            if key in {"normalized_cogs", "cost_rate", "gross_margin_won", "gross_margin_rate"}:
                cell.value = formula_for(row_index, key)
            else:
                cell.value = row_data.get(key)
            style_data_cell(cell, key)
    add_validations(ws, max(2, len(rows) + 1))


def write_base_profiles_sheet(ws, seed: dict[str, Any]) -> None:
    headers = [
        "base_profile_id",
        "label",
        "procurement_type",
        "source_cost_type",
        "branchManagementRate",
        "freightPerUnit",
        "tariffRate",
        "packagingCost",
        "inspectionCost",
        "lossRate",
        "vatRate",
        "includeVatInCogs",
        "shippingPerUnit",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_index, (base_profile_id, profile) in enumerate(seed.get("baseProfiles", {}).items(), start=2):
        fixed = profile.get("fixedDefaults", {})
        ws.cell(row=row_index, column=1, value=base_profile_id)
        ws.cell(row=row_index, column=2, value=profile.get("label", ""))
        ws.cell(row=row_index, column=3, value=profile.get("procurementType", ""))
        ws.cell(row=row_index, column=4, value=profile.get("sourceCostType", ""))
        ws.cell(row=row_index, column=5, value=fixed.get("branchManagementRate"))
        ws.cell(row=row_index, column=6, value=fixed.get("freightPerUnit"))
        ws.cell(row=row_index, column=7, value=fixed.get("tariffRate"))
        ws.cell(row=row_index, column=8, value=fixed.get("packagingCost"))
        ws.cell(row=row_index, column=9, value=fixed.get("inspectionCost"))
        ws.cell(row=row_index, column=10, value=fixed.get("lossRate"))
        ws.cell(row=row_index, column=11, value=fixed.get("vatRate"))
        ws.cell(row=row_index, column=12, value=bool_to_flag(fixed.get("includeVatInCogs")))
        ws.cell(row=row_index, column=13, value=fixed.get("shippingPerUnit"))
    for column in range(1, 14):
        ws.column_dimensions[get_column_letter(column)].width = 18
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:M1"


def write_readme_sheet(ws, output_path: Path) -> None:
    rows = [
        ("PRESSCO21 원가 원장 사용법", True),
        ("1. existing_products 시트에서 기존 상품 원가를 검토합니다.", False),
        ("2. conflict_needs_review 행은 확정원가입력과 검토메모를 채우기 전까지 자동원가로 쓰지 않습니다.", False),
        ("3. new_products 시트는 신규 상품 작성용 템플릿입니다. 맞는 기본프로필 행을 골라 상품명과 원가만 채우면 됩니다.", False),
        ("4. 계산COGS, 원가율, 매출총이익률은 시트에서 자동 계산됩니다.", False),
        ("5. 검토가 끝난 행은 검토상태를 verified_ready로 바꾸고 내보내기=Y로 둡니다.", False),
        ("6. 내보내기 명령:", False),
        (f"   python3 scripts/export-pressco21-cost-ledger.py --input {output_path} --output docs/reference/openclaw-pressco21-product-cost-profiles.reviewed.json", False),
        ("7. reviewed.json이 생기면 가격/마진 에이전트가 그 파일을 seed보다 우선 사용합니다.", False),
        ("8. 기존 연계마스터는 상품 식별용 힌트이고, 이 원장이 확정 원가 관리 기준입니다.", False),
    ]
    ws.column_dimensions["A"].width = 120
    for row_index, (text, is_title) in enumerate(rows, start=1):
        cell = ws.cell(row=row_index, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if is_title:
            cell.font = Font(size=15, bold=True)
            cell.fill = PatternFill("solid", fgColor="E0F2FE")
        elif row_index == 7:
            cell.font = Font(bold=True)
        ws.row_dimensions[row_index].height = 24


def main() -> None:
    args = parse_args()
    seed_path = Path(args.seed).expanduser()
    output_path = Path(args.output).expanduser()
    seed = json.loads(seed_path.read_text(encoding="utf-8"))

    workbook = Workbook()
    readme_ws = workbook.active
    readme_ws.title = "README"
    write_readme_sheet(readme_ws, output_path)

    base_profiles_ws = workbook.create_sheet("base_profiles")
    write_base_profiles_sheet(base_profiles_ws, seed)

    existing_ws = workbook.create_sheet("existing_products")
    write_rows(existing_ws, build_existing_rows(seed))

    new_ws = workbook.create_sheet("new_products")
    write_rows(new_ws, build_new_rows(seed, args.new_rows_per_profile))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Wrote ledger workbook to {output_path}")


if __name__ == "__main__":
    main()
