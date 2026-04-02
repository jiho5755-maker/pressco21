#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_PROFILES = {
    "china_import_floral": {
        "label": "중국 사입 - 플라워/프리저브드 기본형",
        "procurementType": "china_import",
        "sourceCostType": "purchase_cny",
        "fixedDefaults": {
            "branchManagementRate": 0.1,
            "freightPerUnit": 300,
            "tariffRate": 0.13,
            "packagingCost": 150,
            "inspectionCost": 50,
            "lossRate": 0.03,
            "vatRate": 0.1,
            "includeVatInCogs": True,
        },
    },
    "china_import_resin_liquid": {
        "label": "중국 사입 - 레진/용액 기본형",
        "procurementType": "china_import",
        "sourceCostType": "purchase_cny",
        "fixedDefaults": {
            "branchManagementRate": 0.1,
            "freightPerUnit": 400,
            "tariffRate": 0.065,
            "packagingCost": 300,
            "inspectionCost": 100,
            "lossRate": 0.02,
            "vatRate": 0.1,
            "includeVatInCogs": True,
        },
    },
    "china_import_mold_tool": {
        "label": "중국 사입 - 몰드/도구 기본형",
        "procurementType": "china_import",
        "sourceCostType": "purchase_cny",
        "fixedDefaults": {
            "branchManagementRate": 0.1,
            "freightPerUnit": 250,
            "tariffRate": 0.065,
            "packagingCost": 120,
            "inspectionCost": 50,
            "lossRate": 0.02,
            "vatRate": 0.1,
            "includeVatInCogs": True,
        },
    },
    "china_import_fragile_glass": {
        "label": "중국 사입 - 유리/깨짐주의 기본형",
        "procurementType": "china_import",
        "sourceCostType": "purchase_cny",
        "fixedDefaults": {
            "branchManagementRate": 0.1,
            "freightPerUnit": 600,
            "tariffRate": 0.08,
            "packagingCost": 500,
            "inspectionCost": 150,
            "lossRate": 0.03,
            "vatRate": 0.1,
            "includeVatInCogs": True,
        },
    },
    "domestic_purchase_simple": {
        "label": "국내 사입 기본형",
        "procurementType": "domestic_purchase",
        "sourceCostType": "purchase_krw",
        "fixedDefaults": {
            "shippingPerUnit": 100,
            "packagingCost": 100,
            "inspectionCost": 30,
            "lossRate": 0.01,
        },
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate PRESSCO21 draft product cost profiles.")
    parser.add_argument("--source", required=True, help="Path to SKU master workbook")
    parser.add_argument("--output", required=True, help="Path to output JSON")
    parser.add_argument("--top", type=int, default=20, help="Number of grouped products to emit")
    return parser.parse_args()


def split_pipe(value: Any) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).split("|") if str(part).strip()]


def parse_number(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def normalized_join(parts: list[str]) -> str:
    return "|".join(part.strip() for part in parts if part.strip())


def aligned_value(values: list[str], index: int) -> str:
    if not values:
        return ""
    if len(values) == 1:
        return values[0]
    if index < len(values):
        return values[index]
    return ""


def choose_base_profile(name: str, category: str, source_names: list[str]) -> str:
    haystack = " ".join([name, category, *source_names]).lower()
    if category == "레진공예" and any(keyword in haystack for keyword in ["몰드", "실리콘", "도구", "가위", "스티커", "책갈피"]):
        return "china_import_mold_tool"
    if category in ["드라이&프리저브드 플라워", "압화"] and not any(
        keyword in haystack for keyword in ["용액", "코팅액"]
    ):
        return "china_import_floral"
    if any(
        keyword in haystack for keyword in ["보존화", "프리저브드", "수국", "압화", "드라이"]
    ):
        return "china_import_floral"
    if any(keyword in haystack for keyword in ["용액", "코팅액", "착색제", "레진"]):
        return "china_import_resin_liquid"
    if any(keyword in haystack for keyword in ["유리", "돔", "액자", "오르골", "용기", "볼펜케이스"]):
        return "china_import_fragile_glass"
    if category == "레진공예" or any(keyword in haystack for keyword in ["실리콘몰드", "몰드", "도구", "가위", "스티커"]):
        return "china_import_mold_tool"
    return "china_import_simple" if "china_import_simple" in BASE_PROFILES else "china_import_mold_tool"


def build_profiles(source_path: Path, top_n: int) -> dict[str, Any]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    index = {header: position for position, header in enumerate(headers)}

    def cell(row: tuple[Any, ...], header: str) -> Any:
        position = index.get(header)
        if position is None or position >= len(row):
            return None
        return row[position]

    groups = defaultdict(
        lambda: {
            "rows": 0,
            "orderCountSum": 0.0,
            "maxOrderCount": 0.0,
            "latestDate": "",
            "latestCostCny": None,
            "costHistoryCny": set(),
            "priceHistoryKrw": set(),
            "skuCodes": set(),
            "sabangs": set(),
            "sourceNames": set(),
            "category": "",
        }
    )

    for row in rows:
        branduids = split_pipe(cell(row, "메이크샵 branduid"))
        product_names = split_pipe(cell(row, "메이크샵 상품명"))
        if not branduids or not product_names:
            continue
        categories = split_pipe(cell(row, "메이크샵 카테고리"))
        prices = split_pipe(cell(row, "메이크샵 판매가"))
        entry_count = max(len(branduids), len(product_names), len(categories), len(prices))
        for entry_index in range(entry_count):
            branduid = aligned_value(branduids, entry_index)
            product_name = aligned_value(product_names, entry_index)
            if not branduid or not product_name:
                continue
            key = (branduid, product_name)
            group = groups[key]
            group["rows"] += 1
            order_count = parse_number(cell(row, "주문횟수")) or 0.0
            group["orderCountSum"] += order_count
            group["maxOrderCount"] = max(group["maxOrderCount"], order_count)
            latest_date = str(cell(row, "최근주문일") or "").strip()
            latest_cost = parse_number(cell(row, "최근단가(CNY)"))
            if latest_date > group["latestDate"]:
                group["latestDate"] = latest_date
                group["latestCostCny"] = round(latest_cost, 4) if latest_cost is not None else None
            if latest_cost is not None:
                group["costHistoryCny"].add(round(latest_cost, 4))
            numeric_price = parse_number(aligned_value(prices, entry_index))
            if numeric_price is not None:
                group["priceHistoryKrw"].add(int(numeric_price))
            sku_code = str(cell(row, "SKU코드") or "").strip()
            if sku_code:
                group["skuCodes"].add(sku_code)
            sabang_code = str(cell(row, "사방넷 상품코드") or "").strip()
            if sabang_code:
                group["sabangs"].add(sabang_code)
            source_name = str(cell(row, "1688 한글품명") or "").strip()
            if source_name:
                group["sourceNames"].add(source_name)
            if not group["category"]:
                group["category"] = aligned_value(categories, entry_index)

    ranked = []
    for (branduid, name), group in groups.items():
        ranked.append(
            {
                "branduid": branduid,
                "name": name,
                **group,
            }
        )

    ranked.sort(
        key=lambda item: (
            -item["orderCountSum"],
            -item["maxOrderCount"],
            -item["rows"],
            item["name"],
        )
    )

    products = []
    for item in ranked[:top_n]:
        base_profile_id = choose_base_profile(item["name"], item["category"], sorted(item["sourceNames"]))
        cost_history = sorted(item["costHistoryCny"])
        price_history = sorted(item["priceHistoryKrw"])
        latest_purchase_cny = item["latestCostCny"] if item["latestCostCny"] is not None else (cost_history[0] if cost_history else None)
        status = "draft_seeded_from_master"
        notes = [
            "연계마스터/SKU마스터 기반 자동 생성 초안",
            "확정 원가 원장이 아니므로 검증 후 status를 승격해야 함",
        ]
        if len(cost_history) > 1:
            status = "draft_conflict_from_master"
            notes.append(f"동일 상품군 CNY 이력 충돌: {cost_history}")

        products.append(
            {
                "productId": f"pressco21-{item['branduid'].split('|')[0]}",
                "status": status,
                "productName": item["name"],
                "aliases": sorted(item["sourceNames"]),
                "skuCodes": sorted(item["skuCodes"]),
                "branduids": split_pipe(item["branduid"]),
                "sabangCodes": sorted(item["sabangs"]),
                "category": item["category"],
                "baseProfileId": base_profile_id,
                "latestSourceSnapshot": {
                    "latestOrderDate": item["latestDate"],
                    "latestPurchaseCny": latest_purchase_cny,
                    "costHistoryCny": cost_history,
                    "sellingPriceHistoryKrw": price_history,
                    "latestSellingPriceKrw": price_history[0] if len(price_history) == 1 else None,
                    "rowCount": item["rows"],
                    "orderCountSum": item["orderCountSum"],
                    "maxOrderCount": item["maxOrderCount"],
                    "sourceNames": sorted(item["sourceNames"]),
                },
                "fixedOverrides": {},
                "variableInputHints": {
                    "exchangeRate": "최근 적용 환율",
                    "purchaseCny": "최신 발주 단가로 교체 필요",
                },
                "notes": notes,
            }
        )

    return {
        "version": "0.1",
        "updatedAt": "2026-03-21",
        "sourceWorkbook": str(source_path),
        "selectionRule": f"SKU마스터 grouped top {top_n} by orderCountSum",
        "baseProfiles": BASE_PROFILES,
        "products": products,
    }


def main() -> None:
    args = parse_args()
    payload = build_profiles(Path(args.source).expanduser(), args.top)
    output_path = Path(args.output).expanduser()
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(payload['products'])} draft cost profiles to {output_path}")


if __name__ == "__main__":
    main()
