#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_PRODUCT_1688 = "1688 한글품명"
HEADER_SPEC_1688 = "1688 규격"
HEADER_COST_CNY = "최근단가(CNY)"
HEADER_ORDER_DATE = "최근주문일"
HEADER_ORDER_COUNT = "주문횟수"
HEADER_URL_1688 = "1688 URL"
HEADER_DB_CODE = "DB코드"
HEADER_SKU_CODE = "SKU코드"
HEADER_SKU_NAME = "SKU상품명"
HEADER_BRANDUID = "메이크샵 branduid"
HEADER_PRODUCT_NAME = "메이크샵 상품명"
HEADER_SELLING_PRICE = "메이크샵 판매가"
HEADER_CATEGORY = "메이크샵 카테고리"
HEADER_MAP_1688 = "1688→SKU"
HEADER_MAP_MAKESHOP = "SKU→메이크샵"
HEADER_NOTE = "비고"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Look up PRESSCO21 pricing catalog rows from the linked XLSX workbook.",
    )
    parser.add_argument("--catalog", action="append", default=[], help="Path to catalog workbook")
    parser.add_argument("--product-name", default="", help="Product name to match")
    parser.add_argument("--sku", default="", help="SKU code to match")
    parser.add_argument("--branduid", default="", help="MakeShop branduid to match")
    parser.add_argument("--max-candidates", type=int, default=5, help="Number of candidates to return")
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("주식회사", "")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[^0-9a-z가-힣]+", "", text)
    return text


def split_multi(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in re.split(r"\s*\|\s*", text) if part.strip()]


def aligned_value(values: list[str], index: int) -> str | None:
    if not values:
        return None
    if len(values) == 1:
        return values[0]
    if index < len(values):
        return values[index]
    return None


def parse_number(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
    else:
        match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
        if not match:
            return None
        numeric = float(match.group(0))
    if numeric.is_integer():
        return int(numeric)
    return round(numeric, 4)


def compact_number(value: Any) -> str:
    if value is None or value == "":
        return "-"
    return str(value)


@dataclass
class Candidate:
    score: int
    reasons: list[str]
    record: dict[str, Any]


def get_cell(row: tuple[Any, ...], header_index: dict[str, int], header: str) -> Any:
    index = header_index.get(header)
    if index is None or index >= len(row):
        return None
    return row[index]


def load_records(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "연계마스터" if "연계마스터" in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    header_index = {header: position for position, header in enumerate(headers)}
    records: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows, start=2):
        source_product_name = get_cell(row, header_index, HEADER_PRODUCT_1688)
        sku_code = get_cell(row, header_index, HEADER_SKU_CODE)
        sku_product_name = get_cell(row, header_index, HEADER_SKU_NAME)
        if not any([source_product_name, sku_code, sku_product_name]):
            continue

        branduids = split_multi(get_cell(row, header_index, HEADER_BRANDUID))
        product_names = split_multi(get_cell(row, header_index, HEADER_PRODUCT_NAME))
        selling_prices = split_multi(get_cell(row, header_index, HEADER_SELLING_PRICE))
        categories = split_multi(get_cell(row, header_index, HEADER_CATEGORY))
        item_count = max(1, len(branduids), len(product_names), len(selling_prices), len(categories))

        for item_index in range(item_count):
            records.append(
                {
                    "rowNumber": row_number,
                    "sourceProductName": str(source_product_name or "").strip(),
                    "sourceSpec": str(get_cell(row, header_index, HEADER_SPEC_1688) or "").strip(),
                    "recentCostCny": parse_number(get_cell(row, header_index, HEADER_COST_CNY)),
                    "recentOrderDate": get_cell(row, header_index, HEADER_ORDER_DATE),
                    "orderCount": parse_number(get_cell(row, header_index, HEADER_ORDER_COUNT)),
                    "sourceUrl": str(get_cell(row, header_index, HEADER_URL_1688) or "").strip(),
                    "dbCode": str(get_cell(row, header_index, HEADER_DB_CODE) or "").strip(),
                    "skuCode": str(sku_code or "").strip(),
                    "skuProductName": str(sku_product_name or "").strip(),
                    "makeshopBranduid": aligned_value(branduids, item_index),
                    "makeshopProductName": aligned_value(product_names, item_index),
                    "makeshopSellingPrice": parse_number(aligned_value(selling_prices, item_index)),
                    "makeshopCategory": aligned_value(categories, item_index),
                    "mapping1688ToSku": str(get_cell(row, header_index, HEADER_MAP_1688) or "").strip(),
                    "mappingSkuToMakeshop": str(get_cell(row, header_index, HEADER_MAP_MAKESHOP) or "").strip(),
                    "note": str(get_cell(row, header_index, HEADER_NOTE) or "").strip(),
                }
            )

    return records


def score_record(record: dict[str, Any], product_name: str, sku: str, branduid: str) -> Candidate | None:
    normalized_product_name = normalize_text(product_name)
    normalized_sku = normalize_text(sku)
    normalized_branduid = normalize_text(branduid)
    score = 0
    reasons: list[str] = []

    record_sku = normalize_text(record.get("skuCode"))
    record_branduid = normalize_text(record.get("makeshopBranduid"))
    record_product_fields = [
        ("makeshop_exact", normalize_text(record.get("makeshopProductName")), 320, 220),
        ("sku_exact", normalize_text(record.get("skuProductName")), 280, 180),
        ("source_exact", normalize_text(record.get("sourceProductName")), 240, 140),
    ]

    if normalized_sku and record_sku:
        if normalized_sku == record_sku:
            score += 1000
            reasons.append("sku_exact")
        elif normalized_sku in record_sku or record_sku in normalized_sku:
            score += 650
            reasons.append("sku_partial")

    if normalized_branduid and record_branduid:
        if normalized_branduid == record_branduid:
            score += 900
            reasons.append("branduid_exact")
        elif normalized_branduid in record_branduid or record_branduid in normalized_branduid:
            score += 600
            reasons.append("branduid_partial")

    if normalized_product_name:
        for label, candidate, exact_weight, contains_weight in record_product_fields:
            if not candidate:
                continue
            if normalized_product_name == candidate:
                score += exact_weight
                reasons.append(label)
            elif normalized_product_name in candidate or candidate in normalized_product_name:
                score += contains_weight
                reasons.append(label.replace("exact", "contains"))
            ratio = SequenceMatcher(None, normalized_product_name, candidate).ratio()
            score += int(ratio * 70)

    if score <= 0:
        return None

    return Candidate(score=score, reasons=reasons, record=record)


def summarize_record(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "skuCode": record.get("skuCode"),
        "makeshopBranduid": record.get("makeshopBranduid"),
        "productName": record.get("makeshopProductName") or record.get("skuProductName") or record.get("sourceProductName"),
        "recentCostCny": record.get("recentCostCny"),
        "makeshopSellingPrice": record.get("makeshopSellingPrice"),
        "category": record.get("makeshopCategory"),
        "rowNumber": record.get("rowNumber"),
    }


def identity_key(record: dict[str, Any]) -> tuple[str, str, str]:
    return (
        normalize_text(record.get("skuCode")),
        normalize_text(record.get("makeshopBranduid")),
        normalize_text(
            record.get("makeshopProductName")
            or record.get("skuProductName")
            or record.get("sourceProductName")
        ),
    )


def sortable_order_date(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        return datetime.fromisoformat(raw).strftime("%Y%m%d")
    except ValueError:
        return raw


def build_summary(candidate: Candidate, catalog_path: Path) -> str:
    record = candidate.record
    product_name = record.get("makeshopProductName") or record.get("skuProductName") or record.get("sourceProductName")
    return (
        f"{catalog_path.name}에서 매칭 "
        f"(score={candidate.score}, reasons={','.join(candidate.reasons) or '-'}) "
        f"SKU={compact_number(record.get('skuCode'))}, "
        f"branduid={compact_number(record.get('makeshopBranduid'))}, "
        f"상품명={compact_number(product_name)}, "
        f"CNY={compact_number(record.get('recentCostCny'))}, "
        f"메이크샵가={compact_number(record.get('makeshopSellingPrice'))}, "
        f"카테고리={compact_number(record.get('makeshopCategory'))}"
    )


def match_catalog(
    catalog_path: Path,
    records: list[dict[str, Any]],
    product_name: str,
    sku: str,
    branduid: str,
    max_candidates: int,
) -> dict[str, Any]:
    scored = []
    for record in records:
        candidate = score_record(record, product_name, sku, branduid)
        if candidate:
            scored.append(candidate)

    if not scored:
        return {
            "status": "not_found",
            "catalogPath": str(catalog_path),
            "summary": f"{catalog_path.name}에서 일치 항목을 찾지 못했습니다.",
            "candidates": [],
        }

    scored.sort(
        key=lambda item: (
            item.score,
            sortable_order_date(item.record.get("recentOrderDate")),
            item.record.get("rowNumber") or 0,
            1 if item.record.get("makeshopSellingPrice") is not None else 0,
            1 if item.record.get("recentCostCny") is not None else 0,
        ),
        reverse=True,
    )

    top = scored[0]
    second = scored[1] if len(scored) > 1 else None
    exact_key_match = "sku_exact" in top.reasons or "branduid_exact" in top.reasons
    exact_name_match = any(reason.endswith("_exact") for reason in top.reasons)
    safe_gap = second is None or top.score - second.score >= 40
    same_identity_cluster = [
        candidate for candidate in scored if identity_key(candidate.record) == identity_key(top.record)
    ]
    matched = (
        exact_key_match
        or (exact_name_match and safe_gap)
        or (top.score >= 420 and safe_gap)
        or (exact_name_match and len(same_identity_cluster) > 1)
    )
    same_identity_costs = sorted(
        {
            record.record.get("recentCostCny")
            for record in same_identity_cluster
            if record.record.get("recentCostCny") is not None
        }
    )
    trust_status = "stable" if len(same_identity_costs) <= 1 else "conflict"
    trust_summary = (
        "동일 식별자 이력 원가 1개로 안정적"
        if trust_status == "stable"
        else f"동일 식별자 이력 원가가 여러 개라 충돌 가능: {same_identity_costs}"
    )

    if matched:
        return {
            "status": "matched",
            "catalogPath": str(catalog_path),
            "summary": build_summary(top, catalog_path),
            "matchedBy": top.reasons,
            "confidenceScore": top.score,
            "trustStatus": trust_status,
            "trustSummary": trust_summary,
            "sameIdentityCostHistory": same_identity_costs,
            "record": top.record,
            "candidates": [summarize_record(item.record) for item in scored[:max_candidates]],
            "identityHistory": [summarize_record(item.record) for item in same_identity_cluster[:max_candidates]],
        }

    return {
        "status": "ambiguous",
        "catalogPath": str(catalog_path),
        "summary": (
            f"{catalog_path.name}에서 후보가 여러 개라 SKU 또는 branduid가 필요합니다. "
            f"상위 후보 수={min(len(scored), max_candidates)}"
        ),
        "candidates": [summarize_record(item.record) for item in scored[:max_candidates]],
    }


def main() -> None:
    args = parse_args()
    catalog_paths = [Path(path).expanduser() for path in args.catalog]
    if not catalog_paths:
        raise SystemExit("At least one --catalog path is required.")

    product_name = args.product_name.strip()
    sku = args.sku.strip()
    branduid = args.branduid.strip()
    if not any([product_name, sku, branduid]):
        raise SystemExit("At least one of --product-name, --sku, or --branduid is required.")

    for catalog_path in catalog_paths:
        if not catalog_path.exists():
            continue
        records = load_records(catalog_path)
        result = match_catalog(
            catalog_path,
            records,
            product_name,
            sku,
            branduid,
            args.max_candidates,
        )
        if result["status"] == "matched":
            sys.stdout.write(json.dumps(result, ensure_ascii=False, indent=2) + "\n")
            return
        if result["status"] == "ambiguous":
            sys.stdout.write(json.dumps(result, ensure_ascii=False, indent=2) + "\n")
            return

    sys.stdout.write(
        json.dumps(
            {
                "status": "not_found",
                "catalogPath": str(catalog_paths[0]),
                "summary": "지정한 카탈로그들에서 일치 항목을 찾지 못했습니다.",
                "candidates": [],
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )


if __name__ == "__main__":
    main()
