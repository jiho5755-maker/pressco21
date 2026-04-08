#!/usr/bin/env python3
"""
3채널 상품 매칭 엔진 v3 — 공격적 추론 매칭
MakeShop ↔ 스마트스토어/쿠팡 상품 교차 매칭

v2 대비 개선:
1. 상품코드(K101, PF10-2 등) 기반 매칭 → A등급 직행
2. "프레스코21" + SEO 노이즈 단어 제거
3. 슬래시/괄호 정규화로 핵심 단어 추출
4. 쿠팡 가격 0원 시 가격 비교 스킵
5. 유사도 85%+ → B등급 보장
6. 1:N 옵션-변형 감지 (색상/사이즈 변형)
"""

import csv
import json
import os
import re
from difflib import SequenceMatcher

# === 경로 ===
SRC = "/Users/jangjiho/Desktop/사방넷_상품정리"
OUT = f"{SRC}/3채널_통합분석"
API_JSON = f"{SRC}/AI용_CSV/6_원본데이터/전체_API_데이터.json"

SS_FILES = [
    "/Users/jangjiho/Downloads/스마트스토어상품_20260408_071542.xlsx",
    "/Users/jangjiho/Downloads/스마트스토어상품_20260408_071604.xlsx",
]
CP_FILE = "/Users/jangjiho/Downloads/price_inventory_260408.xlsx"

# === 노이즈 제거 ===
NOISE_WORDS = {
    "프레스코21", "프레스코", "pressco21", "pressco",
    "국내", "유일", "국내산", "신상품", "인기",
    "DIY", "diy", "수제", "핸드메이드",
}

NOISE_PREFIXES = [
    r"^\[국내\s*유일\]\s*",
    r"^프레스코21\s*",
    r"^pressco21\s*",
]

# 상품코드 패턴: K101, K304, PF10-2, PF24-8, FP/K228
CODE_PATTERN = re.compile(
    r'(?:FP/?)?(K\d{2,4})'      # K101, K304, FP/K228
    r'|'
    r'(PF\d+-\d+)'              # PF10-2, PF24-8
    r'|'
    r'([A-Z]{1,3}\d{3,5})'      # 기타 3자리+ 숫자 코드
)

# 규격/스펙 패턴: 100g, 5g, 29.5ml, 10ml, 20p
SPEC_PATTERN = re.compile(r'(\d+(?:\.\d+)?)\s*(g|ml|p|mm|cm|oz|개|장|매|세트|인|묶음)', re.I)


def normalize_name(name):
    """상품명 정규화: 노이즈 제거 + 슬래시/괄호 분리"""
    n = name.strip()
    # 접두사 제거
    for pat in NOISE_PREFIXES:
        n = re.sub(pat, '', n, flags=re.I)
    # 노이즈 단어 제거
    for w in NOISE_WORDS:
        n = n.replace(w, '')
    # 슬래시/괄호를 공백으로
    n = re.sub(r'[/\(\)\[\]【】\-–—_]', ' ', n)
    # 다중 공백 → 단일
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def extract_codes(name):
    """상품명에서 코드 추출 (K304, PF10-2 등)"""
    codes = set()
    for m in CODE_PATTERN.finditer(name):
        for g in m.groups():
            if g:
                codes.add(g.upper())
    return codes


def extract_specs(name):
    """규격 추출: {100g, 5ml, 20p}"""
    specs = set()
    for m in SPEC_PATTERN.finditer(name):
        specs.add(f"{m.group(1)}{m.group(2).lower()}")
    return specs


def extract_keywords(name):
    """핵심 키워드 추출 (2글자+ 의미 단어)"""
    n = normalize_name(name)
    # 공통 무의미 단어 제거
    stops = {"만들기", "재료", "공예", "세트", "시리즈", "모음전", "모음",
             "압화", "레진", "아트", "공예재료", "꽃", "북마크", "다꾸",
             "편지", "디자인", "선택", "색상", "포함", "개입", "약",
             "단위", "사이즈", "타입", "기존"}
    words = re.split(r'\s+', n)
    result = []
    for w in words:
        w = w.strip()
        if len(w) >= 2 and w not in stops:
            result.append(w)
    return result


def jaccard_keywords(kw1, kw2):
    """키워드 자카드 유사도 (부분 문자열 포함)"""
    if not kw1 or not kw2:
        return 0
    matches = 0
    total = max(len(kw1), len(kw2))
    for w1 in kw1:
        for w2 in kw2:
            if w1 in w2 or w2 in w1:
                matches += 1
                break
    return matches / total if total else 0


def similarity(name1, name2):
    """종합 유사도 (정규화 후)"""
    n1 = normalize_name(name1)
    n2 = normalize_name(name2)
    return SequenceMatcher(None, n1, n2).ratio()


def match_score_v3(ext_name, ext_price, ms_name, ms_price, ms_uid=None, ext_vcode=None):
    """
    v3 매칭 스코어 (0~100)
    Returns: (score, grade, reasons[])
    """
    reasons = []
    score = 0

    en = normalize_name(ext_name)
    mn = normalize_name(ms_name)

    # 1. 완전 일치 (빠른 탈출)
    if ext_name.strip() == ms_name.strip():
        return 100, "A_확실", ["exact"]
    if en == mn:
        return 98, "A_확실", ["norm_exact"]

    # 2. vcode = uid (쿠팡 업체상품코드)
    if ext_vcode and ms_uid:
        try:
            if str(ext_vcode).strip() == str(ms_uid).strip():
                reasons.append("vcode=uid")
                score += 50
        except Exception:
            pass

    # 3. 상품코드 (K304, PF10-2 등)
    ext_codes = extract_codes(ext_name)
    ms_codes = extract_codes(ms_name)
    code_match = ext_codes & ms_codes
    if code_match:
        reasons.append(f"코드={','.join(code_match)}")
        score += 35

    # 4. 정규화 포함 (핵심 신호!)
    #    짧은 이름이 긴 이름에 완전 포함 → 같은 상품 가능성 높음
    norm_contain = False
    if len(mn) >= 2 and len(en) >= 2:
        if mn in en or en in mn:
            norm_contain = True
            score += 30
            reasons.append("정규화포함")

    # 5. SequenceMatcher 유사도
    sim = SequenceMatcher(None, en, mn).ratio()
    sim_pct = int(sim * 100)
    if sim >= 0.85:
        score += 30
        reasons.append(f"유사{sim_pct}%↑")
    elif sim >= 0.70:
        score += 22
        reasons.append(f"유사{sim_pct}%")
    elif sim >= 0.55:
        score += 14
        reasons.append(f"유사{sim_pct}%")
    elif sim >= 0.40:
        score += 7
        reasons.append(f"유사{sim_pct}%")

    # 6. 키워드 매칭
    ext_kw = extract_keywords(ext_name)
    ms_kw = extract_keywords(ms_name)
    jac = jaccard_keywords(ext_kw, ms_kw)
    if jac >= 0.6:
        score += 12
        reasons.append(f"키워드{int(jac*100)}%")
    elif jac >= 0.4:
        score += 8
        reasons.append(f"키워드{int(jac*100)}%")
    elif jac >= 0.2:
        score += 4
        reasons.append(f"키워드{int(jac*100)}%")

    # 7. 규격 매칭
    ext_specs = extract_specs(ext_name)
    ms_specs = extract_specs(ms_name)
    spec_ok = False
    if ext_specs and ms_specs:
        spec_match = ext_specs & ms_specs
        if spec_match:
            score += 8
            spec_ok = True
            reasons.append(f"규격={','.join(spec_match)}")
        else:
            score -= 5
            reasons.append("규격불일치")

    # 8. 가격 매칭 (0원 스킵)
    try:
        ep = int(float(ext_price)) if ext_price else 0
        mp = int(float(ms_price)) if ms_price else 0
    except (ValueError, TypeError):
        ep, mp = 0, 0

    price_exact = False
    price_close = False
    if ep > 0 and mp > 0:
        if ep == mp:
            price_exact = True
            score += 15
            reasons.append("가격=")
        elif abs(ep - mp) / max(ep, mp) < 0.15:
            price_close = True
            score += 10
            reasons.append("가격≈")
        elif abs(ep - mp) / max(ep, mp) > 0.5:
            score -= 5
            reasons.append("가격×")

    # === 등급 판정 ===
    if score >= 65:
        grade = "A_확실"
    elif score >= 45:
        grade = "B_유력"
    elif score >= 30:
        grade = "C_추정"
    else:
        grade = "D_미매칭"

    # === 보정 규칙 ===

    # 정규화포함 + 가격일치 → A 보장 (v2 핵심 패턴)
    if norm_contain and price_exact and grade != "A_확실":
        grade = "A_확실"
        score = max(score, 80)
        reasons.append("포함+가격→A")

    # 정규화포함 + 가격유사 → 최소 B
    if norm_contain and price_close and grade in ("C_추정", "D_미매칭"):
        grade = "B_유력"
        reasons.append("포함+가격≈→B")

    # 유사도 85%+ → 최소 B
    if sim >= 0.85 and grade in ("C_추정", "D_미매칭"):
        grade = "B_유력"
        reasons.append("유사도보정↑")

    # 코드 매칭 → 최소 B
    if code_match and grade in ("C_추정", "D_미매칭"):
        grade = "B_유력"
        reasons.append("코드보정↑")

    # 코드 매칭 + (가격일치 or 유사도70%+) → A
    if code_match and (price_exact or sim >= 0.70) and grade != "A_확실":
        grade = "A_확실"
        score = max(score, 75)
        reasons.append("코드+검증→A")

    # 유사도 70%+ && 키워드 40%+ → 최소 C
    if sim >= 0.70 and jac >= 0.4 and grade == "D_미매칭":
        grade = "C_추정"
        reasons.append("복합보정↑")

    # 유사도 55%+ && 가격= && 키워드 매칭 → 최소 B
    if sim >= 0.55 and price_exact and jac >= 0.2 and grade in ("C_추정", "D_미매칭"):
        grade = "B_유력"
        reasons.append("유사+가격+키워드→B")

    # 키워드 60%+ && 가격유사 → 최소 C (일본 압화 시리즈 등)
    if jac >= 0.6 and (price_exact or price_close) and grade == "D_미매칭":
        grade = "C_추정"
        reasons.append("키워드+가격→C")

    # 유사도 45%+ && 키워드 50%+ && 가격유사 → 최소 C
    if sim >= 0.45 and jac >= 0.5 and (price_exact or price_close) and grade == "D_미매칭":
        grade = "C_추정"
        reasons.append("복합3→C")

    return score, grade, reasons


def load_makeshop():
    """MakeShop 전체 상품 로드 (강사공간/개인결제 제외)"""
    with open(API_JSON, "r", encoding="utf-8") as f:
        all_products = json.load(f)

    products = []
    for p in all_products:
        # JSON은 api_product_audit.py가 만든 것 → cate1 필드 있음
        cate1 = p.get("cate1", "")
        if not cate1:
            cat = p.get("category", "")
            cate1 = cat.split(" < ")[0] if " < " in cat else cat
        name = p.get("product_name", "")
        if cate1 == "강사공간":
            continue
        if cate1 in ["개인결제창", "개인결제"] or "개인결제" in name:
            continue
        products.append({
            "uid": str(p.get("uid", "")),
            "name": name,
            "price": str(p.get("sellprice", "0")),
            "category": cate1,
            "display": p.get("display", "N"),
        })
    return products


def load_smartstore():
    """스마트스토어 상품 로드"""
    import openpyxl
    products = []
    seen_pids = set()

    for fpath in SS_FILES:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active

        # 헤더 찾기 (2번째 행이 실제 필드명)
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            cells = [str(c).strip() if c else "" for c in row]
            if "상품번호" in cells or "판매가" in cells:
                headers = cells
                header_row = list(ws.iter_rows(min_row=1, max_row=5, values_only=True)).index(tuple(row)) + 1
                break

        if not headers:
            # 2번째 행이 헤더인 패턴
            row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            headers = [str(c).strip() if c else "" for c in row2]
            header_row = 2

        # 컬럼 인덱스
        col_map = {}
        for i, h in enumerate(headers):
            if "상품번호" in h:
                col_map["pid"] = i
            elif h == "판매자 상품코드":
                col_map["seller_code"] = i
            elif h == "상품명":
                col_map["name"] = i
            elif h == "판매가":
                col_map["price"] = i
            elif "판매상태" in h:
                col_map["status"] = i

        if "name" not in col_map:
            continue

        # 가이드 행 건너뛰기: 실제 데이터는 상품번호가 숫자인 행부터
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cells = list(row)
            # 상품번호가 숫자인 행만 처리 (가이드/설명 행 제외)
            pid_raw = str(cells[col_map.get("pid", 0)]).strip() if col_map.get("pid") is not None and cells[col_map.get("pid", 0)] else ""
            if not pid_raw.isdigit():
                continue

            name = str(cells[col_map["name"]]).strip() if col_map.get("name") is not None and cells[col_map["name"]] else ""
            if not name or name == "None":
                continue
            pid = pid_raw
            price = str(cells[col_map.get("price", 0)]).strip() if col_map.get("price") is not None and cells[col_map.get("price", 0)] else "0"

            if pid in seen_pids:
                continue
            seen_pids.add(pid)

            seller_code = str(cells[col_map.get("seller_code", 0)]).strip() if col_map.get("seller_code") is not None and col_map["seller_code"] < len(cells) and cells[col_map["seller_code"]] else ""
            if seller_code == "None":
                seller_code = ""

            products.append({
                "pid": pid,
                "name": name,
                "price": price.replace(",", "").replace("원", "").strip(),
                "seller_code": seller_code,
            })
        wb.close()

    return products


def load_coupang():
    """쿠팡 상품 로드 (옵션 행 합치기)"""
    import openpyxl
    wb = openpyxl.load_workbook(CP_FILE, data_only=True)
    ws = wb.active

    # 헤더: 3행 (1=안내, 2=조회결과, 3=실제 헤더)
    row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    headers = [str(c).strip() if c else "" for c in row3]
    header_row = 3

    col_map = {}
    for i, h in enumerate(headers):
        if "쿠팡" in h and "상품명" in h:
            col_map["name"] = i
        elif "등록" in h and "상품명" in h:
            col_map["reg_name"] = i
        elif h == "업체상품 ID" or h == "업체상품ID":
            col_map["vid"] = i
        elif h == "Product ID":
            col_map["product_id"] = i
        elif "승인상태" in h:
            col_map["approval"] = i
        elif "업체상품코드" in h:
            col_map["vcode"] = i
        elif "판매가격" in h:
            col_map["price"] = i

    print(f"  쿠팡 헤더: {headers[:10]}")
    print(f"  컬럼맵: {col_map}")

    # 업체상품ID 기준 그룹핑
    products = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cells = list(row)
        if not cells or all(c is None for c in cells):
            continue

        def get_cell(key, default=""):
            idx = col_map.get(key)
            if idx is None or idx >= len(cells) or cells[idx] is None:
                return default
            return str(cells[idx]).strip()

        vid = get_cell("vid")
        if not vid or vid == "None":
            continue

        name = get_cell("name")
        reg_name = get_cell("reg_name")
        approval = get_cell("approval")
        vcode = get_cell("vcode")
        product_id = get_cell("product_id")
        price = get_cell("price", "0").replace(",", "").replace("원", "").strip()

        if not name or name == "None":
            name = reg_name
        if not name or name == "None":
            continue

        if vid not in products:
            products[vid] = {
                "vid": vid,
                "product_id": product_id,
                "name": name,
                "reg_name": reg_name,
                "approval": approval,
                "vcode": vcode,
                "price": price,
            }

    wb.close()
    return list(products.values())


def find_best_match(ext_product, ms_products, ms_norm_map, ms_name_set, is_coupang=False):
    """외부 채널 상품 1개에 대해 최적 MakeShop 매칭 (전수비교)"""
    ext_name = ext_product["name"]
    ext_price = ext_product.get("price", "0")
    ext_vcode = ext_product.get("vcode", "") if is_coupang else None

    # 빠른 경로: 완전 일치
    if ext_name.strip() in ms_name_set:
        for p in ms_products:
            if p["name"] == ext_name.strip():
                return 100, "A_확실", ["exact"], p

    nm = normalize_name(ext_name)
    if nm in ms_norm_map:
        return 98, "A_확실", ["norm_exact"], ms_norm_map[nm]

    # 사전 계산
    ext_codes = extract_codes(ext_name)
    ext_kw = extract_keywords(ext_name)
    ext_specs = extract_specs(ext_name)

    best_score = 0
    best_grade = "D_미매칭"
    best_reasons = []
    best_ms = None

    names_to_try = [(ext_name, ext_price, False)]
    if is_coupang and ext_product.get("reg_name") and ext_product["reg_name"] != ext_name:
        names_to_try.append((ext_product["reg_name"], ext_price, True))

    for try_name, try_price, is_alt in names_to_try:
        for ms in ms_products:
            score, grade, reasons = match_score_v3(
                try_name, try_price,
                ms["name"], ms["price"],
                ms["uid"], ext_vcode
            )
            if is_alt and score > best_score:
                reasons = reasons + ["(등록명)"]
            if score > best_score:
                best_score = score
                best_grade = grade
                best_reasons = reasons
                best_ms = ms

    return best_score, best_grade, best_reasons, best_ms


def write_csv(path, rows, fieldnames):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    os.makedirs(OUT, exist_ok=True)

    print("=== 3채널 매칭 엔진 v3 ===\n")

    # 데이터 로드
    print("MakeShop 로드...")
    ms_products = load_makeshop()
    print(f"  → {len(ms_products)}개\n")

    print("스마트스토어 로드...")
    ss_products = load_smartstore()
    print(f"  → {len(ss_products)}개\n")

    print("쿠팡 로드...")
    cp_products = load_coupang()
    print(f"  → {len(cp_products)}개\n")

    # 빠른 조회용 맵
    ms_norm_map = {}
    ms_name_set = set()
    for p in ms_products:
        ms_name_set.add(p["name"])
        nm = normalize_name(p["name"])
        ms_norm_map[nm] = p

    # === 스마트스토어 매칭 ===
    print("스마트스토어 매칭 중...")
    ss_results = []
    ss_grades = {"A_확실": 0, "B_유력": 0, "C_추정": 0, "D_미매칭": 0}

    for i, ss in enumerate(ss_products):
        score, grade, reasons, ms = find_best_match(ss, ms_products, ms_norm_map, ms_name_set)
        ss_grades[grade] = ss_grades.get(grade, 0) + 1

        ms_uid = ms["uid"] if ms else ""
        ss_results.append({
            "등급": grade,
            "score": score,
            "reason": "+".join(reasons) if reasons else "",
            "ss_상품명": ss["name"],
            "ms_상품명": ms["name"] if ms else "",
            "ss_가격": ss["price"],
            "ms_가격": ms["price"] if ms else "",
            "ss_상품번호": ss["pid"],
            "ss_판매자코드": ss.get("seller_code", ""),
            "ms_uid": ms_uid,
            "ms_링크": f"https://www.foreverlove.co.kr/shop/shopdetail.html?branduid={ms_uid}" if ms_uid else "",
        })

        if (i + 1) % 200 == 0:
            print(f"  {i+1}/{len(ss_products)}")

    ss_results.sort(key=lambda x: (-x["score"], x["등급"]))
    ss_fields = ["등급", "score", "reason", "ss_상품명", "ms_상품명", "ss_가격", "ms_가격", "ss_상품번호", "ss_판매자코드", "ms_uid", "ms_링크"]
    write_csv(f"{OUT}/v3_매칭_스마트스토어_{len(ss_products)}건.csv", ss_results, ss_fields)

    print(f"\n  스마트스토어 결과:")
    for g in ["A_확실", "B_유력", "C_추정", "D_미매칭"]:
        print(f"    {g}: {ss_grades[g]}건")

    # === 쿠팡 매칭 ===
    print("\n쿠팡 매칭 중...")
    cp_results = []
    cp_grades = {"A_확실": 0, "B_유력": 0, "C_추정": 0, "D_미매칭": 0}

    for i, cp in enumerate(cp_products):
        score, grade, reasons, ms = find_best_match(cp, ms_products, ms_norm_map, ms_name_set, is_coupang=True)
        cp_grades[grade] = cp_grades.get(grade, 0) + 1

        ms_uid = ms["uid"] if ms else ""
        cp_results.append({
            "등급": grade,
            "score": score,
            "reason": "+".join(reasons) if reasons else "",
            "cp_상품명": cp["name"],
            "ms_상품명": ms["name"] if ms else "",
            "cp_가격": cp["price"],
            "ms_가격": ms["price"] if ms else "",
            "cp_업체상품ID": cp["vid"],
            "cp_ProductID": cp.get("product_id", ""),
            "cp_업체상품코드": cp.get("vcode", ""),
            "cp_승인상태": cp.get("approval", ""),
            "ms_uid": ms_uid,
            "ms_링크": f"https://www.foreverlove.co.kr/shop/shopdetail.html?branduid={ms_uid}" if ms_uid else "",
        })

        if (i + 1) % 200 == 0:
            print(f"  {i+1}/{len(cp_products)}")

    cp_results.sort(key=lambda x: (-x["score"], x["등급"]))
    cp_fields = ["등급", "score", "reason", "cp_상품명", "ms_상품명", "cp_가격", "ms_가격", "cp_업체상품ID", "cp_ProductID", "cp_업체상품코드", "cp_승인상태", "ms_uid", "ms_링크"]
    write_csv(f"{OUT}/v3_매칭_쿠팡_{len(cp_products)}건.csv", cp_results, cp_fields)

    print(f"\n  쿠팡 결과:")
    for g in ["A_확실", "B_유력", "C_추정", "D_미매칭"]:
        print(f"    {g}: {cp_grades[g]}건")

    # === 채널 분포 ===
    print("\n채널 분포 계산...")
    ms_in_ss = set()
    ms_in_cp = set()

    for r in ss_results:
        if r["등급"] in ("A_확실", "B_유력") and r["ms_uid"]:
            ms_in_ss.add(r["ms_uid"])
    for r in cp_results:
        if r["등급"] in ("A_확실", "B_유력") and r["ms_uid"]:
            ms_in_cp.add(r["ms_uid"])

    both = ms_in_ss & ms_in_cp
    ss_only = ms_in_ss - ms_in_cp
    cp_only = ms_in_cp - ms_in_ss
    ms_all_uids = {p["uid"] for p in ms_products}
    ms_only = ms_all_uids - ms_in_ss - ms_in_cp

    print(f"\n  === 채널 분포 (B+등급 신뢰) ===")
    print(f"  3채널 모두: {len(both)}개")
    print(f"  MS+SS만:    {len(ss_only)}개")
    print(f"  MS+CP만:    {len(cp_only)}개")
    print(f"  MS만:       {len(ms_only)}개")

    # 통합 CSV
    dist_rows = []
    for p in ms_products:
        uid = p["uid"]
        channels = []
        if uid in ms_in_ss:
            channels.append("SS")
        if uid in ms_in_cp:
            channels.append("CP")
        channels.insert(0, "MS")
        dist_rows.append({
            "uid": uid,
            "name": p["name"],
            "price": p["price"],
            "category": p["category"],
            "channels": "+".join(channels),
            "channel_count": len(channels),
            "in_ss": "Y" if uid in ms_in_ss else "",
            "in_cp": "Y" if uid in ms_in_cp else "",
        })
    dist_rows.sort(key=lambda x: (-x["channel_count"], x["name"]))
    write_csv(
        f"{OUT}/v3_통합_채널분포_{len(dist_rows)}건.csv",
        dist_rows,
        ["uid", "name", "price", "category", "channels", "channel_count", "in_ss", "in_cp"]
    )

    # === v2 vs v3 비교 ===
    print(f"\n{'='*50}")
    print(f"v3 총 결과 요약")
    print(f"{'='*50}")
    print(f"\n스마트스토어 ({len(ss_products)}건):")
    print(f"  A={ss_grades['A_확실']} B={ss_grades['B_유력']} C={ss_grades['C_추정']} D={ss_grades['D_미매칭']}")
    print(f"  A+B 매칭률: {(ss_grades['A_확실']+ss_grades['B_유력'])/len(ss_products)*100:.1f}%")
    print(f"\n쿠팡 ({len(cp_products)}건):")
    print(f"  A={cp_grades['A_확실']} B={cp_grades['B_유력']} C={cp_grades['C_추정']} D={cp_grades['D_미매칭']}")
    if cp_products:
        print(f"  A+B 매칭률: {(cp_grades['A_확실']+cp_grades['B_유력'])/len(cp_products)*100:.1f}%")
    print(f"\n저장: {OUT}/v3_*.csv")


if __name__ == "__main__":
    main()
