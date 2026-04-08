#!/usr/bin/env python3
"""v3 매칭 결과 → 대표용 XLSX 생성"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/Users/jangjiho/Desktop/사방넷_상품정리/3채널_통합분석"
DST = "/Users/jangjiho/Desktop/사방넷_상품정리/대표용_엑셀"

GRADE_FILLS = {
    "A_확실": PatternFill("solid", fgColor="C6EFCE"),  # 녹색
    "B_유력": PatternFill("solid", fgColor="DDEBF7"),  # 파랑
    "C_추정": PatternFill("solid", fgColor="FFF2CC"),  # 노랑
    "D_미매칭": PatternFill("solid", fgColor="FCE4EC"),  # 분홍
}
HEADER_FILL = PatternFill("solid", fgColor="2B3D2F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", fgColor="F5F7F4")
BORDER = Border(bottom=Side(style="thin", color="E0E0E0"))


def csv_to_sheet(wb, sheet_name, csv_path, link_col_name=None, grade_col="등급"):
    if not os.path.exists(csv_path):
        return 0
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if len(rows) < 2:
        return 0

    ws = wb.create_sheet(sheet_name)
    headers = rows[0]

    # 헤더
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    grade_idx = headers.index(grade_col) if grade_col in headers else None

    # 링크 컬럼 찾기
    link_cols = set()
    for i, h in enumerate(headers):
        if "링크" in h or "link" in h.lower():
            link_cols.add(i)

    LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")

    for row_idx, row in enumerate(rows[1:], 2):
        grade = row[grade_idx] if grade_idx is not None and grade_idx < len(row) else ""
        grade_fill = GRADE_FILLS.get(grade)

        for col_idx, val in enumerate(row):
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=val)
            cell.font = DATA_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if grade_fill and col_idx == grade_idx:
                cell.fill = grade_fill
            elif row_idx % 2 == 0:
                cell.fill = ALT_FILL
            # 링크 컬럼: 하이퍼링크 처리
            if col_idx in link_cols and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = LINK_FONT
                cell.value = "열기"

    # 열 너비
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row in rows[1:min(30, len(rows))]:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        width = min(max_len + 3, 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)}"
    ws.freeze_panes = "A2"
    return len(rows) - 1


def make_summary(wb, ss_grades, cp_grades, dist_stats):
    ws = wb.create_sheet("요약", 0)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15

    data = [
        ["3채널 상품 매칭 v3", "", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["", "A 확실", "B 유력", "C 추정", "D 미매칭", "A+B 매칭률"],
        ["스마트스토어", ss_grades.get("A_확실", 0), ss_grades.get("B_유력", 0),
         ss_grades.get("C_추정", 0), ss_grades.get("D_미매칭", 0), ""],
        ["쿠팡", cp_grades.get("A_확실", 0), cp_grades.get("B_유력", 0),
         cp_grades.get("C_추정", 0), cp_grades.get("D_미매칭", 0), ""],
        ["", "", "", "", "", ""],
        ["채널 분포 (B+등급 기준)", "", "", "", "", ""],
        ["3채널 모두 (MS+SS+CP)", dist_stats["both"], "", "", "", ""],
        ["MS+SS만", dist_stats["ss_only"], "", "", "", ""],
        ["MS+CP만", dist_stats["cp_only"], "", "", "", ""],
        ["MS만", dist_stats["ms_only"], "", "", "", ""],
        ["", "", "", "", "", ""],
        ["등급 설명", "", "", "", "", ""],
        ["A: 이름 일치 + 가격 일치 → 동일 상품 확실", "", "", "", "", ""],
        ["B: 이름 유사 + 코드/가격 일부 일치 → 동일 상품 유력", "", "", "", "", ""],
        ["C: 키워드/유사도 일부 일치 → 같은 상품 추정", "", "", "", "", ""],
        ["D: 매칭 실패 → 수동 확인 필요", "", "", "", "", ""],
    ]

    for r, row in enumerate(data, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = DATA_FONT
            if r == 1:
                cell.font = Font(name="Arial", bold=True, size=14, color="2B3D2F")
            if r == 3:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            if r in (4, 5) and c >= 2 and c <= 5:
                if c == 2:
                    cell.fill = GRADE_FILLS["A_확실"]
                elif c == 3:
                    cell.fill = GRADE_FILLS["B_유력"]
                elif c == 4:
                    cell.fill = GRADE_FILLS["C_추정"]
                elif c == 5:
                    cell.fill = GRADE_FILLS["D_미매칭"]

    # A+B 비율 수식
    ws.cell(row=4, column=6, value="=SUM(B4:C4)/SUM(B4:E4)").number_format = "0.0%"
    ws.cell(row=5, column=6, value="=SUM(B5:C5)/SUM(B5:E5)").number_format = "0.0%"


def main():
    os.makedirs(DST, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # 데이터 로드 + 등급 카운트
    ss_grades = {}
    cp_grades = {}

    ss_path = None
    cp_path = None
    for f in os.listdir(SRC):
        if f.startswith("v3_매칭_스마트스토어") and f.endswith(".csv"):
            ss_path = f"{SRC}/{f}"
        elif f.startswith("v3_매칭_쿠팡") and f.endswith(".csv"):
            cp_path = f"{SRC}/{f}"

    if ss_path:
        with open(ss_path, "r", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if row and row[0] in GRADE_FILLS:
                    ss_grades[row[0]] = ss_grades.get(row[0], 0) + 1

    if cp_path:
        with open(cp_path, "r", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if row and row[0] in GRADE_FILLS:
                    cp_grades[row[0]] = cp_grades.get(row[0], 0) + 1

    # 채널 분포
    dist_path = None
    for f in os.listdir(SRC):
        if f.startswith("v3_통합_채널분포") and f.endswith(".csv"):
            dist_path = f"{SRC}/{f}"

    dist_stats = {"both": 0, "ss_only": 0, "cp_only": 0, "ms_only": 0}
    if dist_path:
        with open(dist_path, "r", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                ch = row.get("channels", "")
                if "SS" in ch and "CP" in ch:
                    dist_stats["both"] += 1
                elif "SS" in ch:
                    dist_stats["ss_only"] += 1
                elif "CP" in ch:
                    dist_stats["cp_only"] += 1
                else:
                    dist_stats["ms_only"] += 1

    # 요약 시트
    make_summary(wb, ss_grades, cp_grades, dist_stats)

    # 매칭 결과 시트
    if ss_path:
        n = csv_to_sheet(wb, "SS매칭", ss_path)
        print(f"  SS매칭: {n}건")
    if cp_path:
        n = csv_to_sheet(wb, "CP매칭", cp_path)
        print(f"  CP매칭: {n}건")

    # 채널 분포 시트
    if dist_path:
        n = csv_to_sheet(wb, "채널분포", dist_path, grade_col="channels")
        print(f"  채널분포: {n}건")

    out = f"{DST}/6_3채널매칭결과_v3.xlsx"
    wb.save(out)
    print(f"\n완료: {out}")


if __name__ == "__main__":
    main()
