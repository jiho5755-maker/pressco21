#!/usr/bin/env python3
"""CSV → XLSX 변환 + 보기 좋게 포맷팅"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/Users/jangjiho/Desktop/사방넷_상품정리/AI용_CSV"
DST = "/Users/jangjiho/Desktop/사방넷_상품정리/대표용_엑셀"

HEADER_FILL = PatternFill("solid", fgColor="2B3D2F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
ALERT_FILL = PatternFill("solid", fgColor="FFF2CC")
BORDER = Border(
    bottom=Side(style="thin", color="E0E0E0"),
)
ALT_FILL = PatternFill("solid", fgColor="F5F7F4")


def csv_to_sheet(wb, sheet_name, csv_path, highlight_cols=None):
    """CSV를 워크시트로 변환"""
    if not os.path.exists(csv_path):
        return
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if len(rows) < 2:
        return

    ws = wb.create_sheet(sheet_name)
    headers = rows[0]

    # 헤더
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 데이터
    link_col = None
    for i, h in enumerate(headers):
        if h in ["링크", "url", "product_url"]:
            link_col = i

    for row_idx, row in enumerate(rows[1:], 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row):
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=val)
            cell.font = DATA_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if fill:
                cell.fill = fill
            # 링크 컬럼
            if col_idx == link_col and val.startswith("http"):
                cell.font = LINK_FONT
                cell.hyperlink = val
                cell.value = "열기"
            # 하이라이트 컬럼
            if highlight_cols and headers[col_idx] in highlight_cols:
                cell.fill = ALERT_FILL

    # 열 너비 자동 조절
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row in rows[1:min(50, len(rows))]:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        width = min(max_len + 4, 50)
        if h in ["링크", "url", "product_url"]:
            width = 8
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 필터
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)}"
    # 틀 고정
    ws.freeze_panes = "A2"

    return len(rows) - 1


def make_품절정리():
    wb = Workbook()
    wb.remove(wb.active)

    n1 = csv_to_sheet(wb, "전체품절_진열중", f"{SRC}/2_품절정리/1_전체품절_진열중_201건.csv",
                       highlight_cols=["판단"])
    n2 = csv_to_sheet(wb, "부분품절", f"{SRC}/2_품절정리/2_부분품절_26건.csv",
                       highlight_cols=["품절옵션", "판단"])
    n3 = csv_to_sheet(wb, "미진열_판매가능", f"{SRC}/2_품절정리/3_미진열_판매가능_106건.csv",
                       highlight_cols=["판단"])

    # 요약 시트
    ws = wb.create_sheet("요약", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 50

    summary = [
        ["품절 정리 현황", "", ""],
        ["", "", ""],
        ["분류", "건수", "할 일"],
        ["전체품절 + 진열 중", n1 or 0, "입고됐으면 품절해제, 아니면 미진열"],
        ["부분품절 (일부 옵션만)", n2 or 0, "입고된 옵션만 품절해제 ← 1순위"],
        ["미진열 + 판매가능", n3 or 0, "팔 거면 진열, 안 팔 거면 그대로"],
    ]
    for r, row in enumerate(summary, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = DATA_FONT
            if r == 1:
                cell.font = Font(name="Arial", bold=True, size=14, color="2B3D2F")
            if r == 3:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            if r == 5:
                cell.fill = ALERT_FILL

    out = f"{DST}/1_품절정리.xlsx"
    wb.save(out)
    print(f"  {out} ({(n1 or 0)+(n2 or 0)+(n3 or 0)}건)")


def make_사방넷등록():
    wb = Workbook()
    wb.remove(wb.active)

    counts = {}
    files = [
        ("단품_2120건", "1_단품_2120건.csv"),
        ("옵션1단_290건", "2_옵션1단_290건.csv"),
        ("옵션2단_38건", "3_옵션2단_38건.csv"),
        ("옵션3단이상_9건", "4_옵션3단이상_9건.csv"),
        ("추가상품만_18건", "5_추가상품만_18건.csv"),
        ("추가상품포함_75건", "6_추가상품포함_75건_중복.csv"),
    ]
    for sheet_name, fname in files:
        n = csv_to_sheet(wb, sheet_name, f"{SRC}/1_사방넷등록용/{fname}",
                         highlight_cols=["select_names", "addition_names", "select_details"])
        counts[sheet_name] = n or 0

    # 요약 시트
    ws = wb.create_sheet("요약", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40

    rows = [
        ["사방넷 등록용 상품 분류", "", ""],
        ["", "", ""],
        ["분류", "건수", "등록 난이도"],
        ["단품 (옵션 없음)", counts.get("단품_2120건", 0), "바로 등록 가능"],
        ["SELECT 옵션 1단", counts.get("옵션1단_290건", 0), "옵션 매핑 후 등록"],
        ["SELECT 옵션 2단", counts.get("옵션2단_38건", 0), "조합 옵션 설정"],
        ["SELECT 옵션 3단+", counts.get("옵션3단이상_9건", 0), "복잡, 수동 작업"],
        ["ADDITION만", counts.get("추가상품만_18건", 0), "추가상품 등록"],
        ["추가상품 포함 (중복)", counts.get("추가상품포함_75건", 0), "다른 시트와 중복, 참고용"],
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = DATA_FONT
            if r == 1:
                cell.font = Font(name="Arial", bold=True, size=14, color="2B3D2F")
            if r == 3:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            if r == 4:
                cell.fill = PatternFill("solid", fgColor="D5E8D4")

    out = f"{DST}/2_사방넷등록용.xlsx"
    wb.save(out)
    print(f"  {out}")


def make_이미지호스팅():
    wb = Workbook()
    wb.remove(wb.active)
    for fname in sorted(os.listdir(f"{SRC}/3_이미지호스팅")):
        if not fname.endswith(".csv"):
            continue
        host = fname.replace("이미지호스트_", "").replace(".csv", "")
        short = host[:31]  # 시트명 31자 제한
        csv_to_sheet(wb, short, f"{SRC}/3_이미지호스팅/{fname}")

    out = f"{DST}/3_이미지호스팅.xlsx"
    wb.save(out)
    print(f"  {out}")


def make_기타이슈():
    wb = Workbook()
    wb.remove(wb.active)
    csv_to_sheet(wb, "가격0원_136건", f"{SRC}/4_기타이슈/가격0원_136건.csv")
    csv_to_sheet(wb, "이미지없음_29건", f"{SRC}/4_기타이슈/이미지없음_29건.csv")
    csv_to_sheet(wb, "미진열전체_578건", f"{SRC}/4_기타이슈/미진열전체_578건.csv")

    out = f"{DST}/4_기타이슈.xlsx"
    wb.save(out)
    print(f"  {out}")


def main():
    os.makedirs(DST, exist_ok=True)
    print("XLSX 생성 중...")
    make_품절정리()
    make_사방넷등록()
    make_이미지호스팅()
    make_기타이슈()
    print(f"\n완료: {DST}/")


if __name__ == "__main__":
    main()
